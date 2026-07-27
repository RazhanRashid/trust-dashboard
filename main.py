"""main.py — Trust Level Dashboard (PyQt6 / cool slate)

Drop-in replacement for the warm-aesthetic main.py. Same analyzer code,
same threading model, same session persistence — only the UI layer is
rewritten on top of PyQt6 + pyqtgraph.

Run (Python 3.10+, developed on 3.14):
    source .venv/bin/activate
    pip install -r requirements.txt          # PyQt6 + pyqtgraph
    python main.py
"""

import os
import subprocess
import sys
import openpyxl
import json
import math
import time
import signal
import threading
import logging
from datetime import datetime
from pathlib import Path

logging.getLogger("root").setLevel(logging.ERROR)

import cv2
import numpy as np
import sounddevice as sd

# OpenCV capture backend — AVFoundation only exists on macOS; DirectShow is the
# reliable low-latency choice on Windows; CAP_ANY lets OpenCV pick on Linux.
if sys.platform == "darwin":
    _CAM_BACKEND = cv2.CAP_AVFOUNDATION
elif sys.platform == "win32":
    _CAM_BACKEND = cv2.CAP_DSHOW
else:
    _CAM_BACKEND = cv2.CAP_ANY

from PyQt6.QtCore import Qt, QTimer, QUrl
from PyQt6.QtGui import QFont, QDesktopServices
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QStackedWidget,
                              QVBoxLayout, QHBoxLayout, QGridLayout, QMessageBox,
                              QFileDialog)

# ── Analyzer modules (unchanged) ─────────────────────────────────────────────
from Physio_analysis.face_analyzer    import FaceAnalyzer
from Physio_analysis.vocal_analyzer   import VocalAnalyzer
from Physio_analysis.trust_engine     import TrustEngine, SCORE_CONFIG, SCORE_VERSION
from Physio_analysis.workload_engine  import WorkloadEngine
from Physio_analysis.hrv_analyzer     import HRVAnalyzer

# ── UI modules ───────────────────────────────────────────────────────────────
from theme import (BG, BG_DEEP, PANEL, LINE, LINE_SOFT, TEXT, TEXT_FAINT, TEXT_GHOST,
                    ACCENT, DANGER, C_WORKLOAD,
                    ui_font, load_packaged_fonts, trust_band, TRUST_BANDS,
                    init_ui_scale, sp)
from panels import (TopStrip, CameraPanel, ScorePanel, VoicePanel, HistoryChart,
                     Footer, FlagSidebar, BlendshapeWatch)
from overlays import OverviewScreen, CalibrationOverlay, SessionSummary
from demographics_dialog import DemographicsDialog
from camera_dialog import CameraDialog
import camera_scanner

try:
    import websockets
    import asyncio as _asyncio
    _HAS_WS = True
except ImportError:
    _HAS_WS = False

# Canonical MediaPipe blendshape names, ARKit order (index 0 = _neutral).
# Shared by the Excel export column order (_build_excel) and the live
# Blendshape Watch panel's selector, so both stay in sync with one list.
BLENDSHAPE_NAMES = [
    "_neutral",
    "browDownLeft",    "browDownRight",    "browInnerUp",
    "browOuterUpLeft", "browOuterUpRight",
    "cheekPuff",       "cheekSquintLeft",  "cheekSquintRight",
    "eyeBlinkLeft",    "eyeBlinkRight",
    "eyeLookDownLeft", "eyeLookDownRight",
    "eyeLookInLeft",   "eyeLookInRight",
    "eyeLookOutLeft",  "eyeLookOutRight",
    "eyeLookUpLeft",   "eyeLookUpRight",
    "eyeSquintLeft",   "eyeSquintRight",
    "eyeWideLeft",     "eyeWideRight",
    "jawForward",      "jawLeft",          "jawOpen",         "jawRight",
    "mouthClose",
    "mouthDimpleLeft", "mouthDimpleRight",
    "mouthFrownLeft",  "mouthFrownRight",
    "mouthFunnel",     "mouthLeft",
    "mouthLowerDownLeft", "mouthLowerDownRight",
    "mouthPressLeft",  "mouthPressRight",
    "mouthPucker",     "mouthRight",
    "mouthRollLower",  "mouthRollUpper",
    "mouthShrugLower", "mouthShrugUpper",
    "mouthSmileLeft",  "mouthSmileRight",
    "mouthStretchLeft","mouthStretchRight",
    "mouthUpperUpLeft","mouthUpperUpRight",
    "noseSneerLeft",   "noseSneerRight",
]


# ═══════════════════════════════════════════════════════════════════════════
class TrustDashboard(QMainWindow):
    """Top-level window. Hosts a QStackedWidget that swaps between
    Overview → Calibration → Live → Waiting → Summary."""

    CAM_W, CAM_H = 320, 240

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Trust Level Dashboard")
        self.setStyleSheet(f"background: {BG};")
        # Scaled like everything else, so the minimum can't demand more room
        # than the screen the design was just scaled down to fit.
        self.resize(sp(1400), sp(980))
        self.setMinimumSize(sp(1200), sp(880))

        # ── Analyzers (same instances as before) ─────────────────────────────
        self.face     = FaceAnalyzer()
        self.vocal    = VocalAnalyzer()
        self.trust    = TrustEngine()
        self.workload = WorkloadEngine()
        self.hrv      = HRVAnalyzer()

        # ── Threading + shared state ─────────────────────────────────────────
        self._lock = threading.Lock()
        self._pending_frame = None
        self._last_frame    = None   # (frame_bgr, face_data)
        self._last_vocal    = None
        self._audio_buffer  = np.zeros(4096)
        self._sample_rate   = 44100
        self._running       = True
        self._cap = None
        # Guards the capture handle itself. The camera loop blocks inside
        # cap.read(); without this a swap could release the device out from
        # under a read in progress, which is a crash rather than a dropped frame.
        self._cap_lock = threading.Lock()
        self._measured_cam_fps: float = 30.0   # set for real in _start_camera / _switch_camera

        # ── Session / history state ─────────────────────────────────────────
        self._history = {k: [] for k in ("total", "facial", "vocal", "gaze", "hrv")}
        self._history_t: list[float] = []   # elapsed-session seconds, parallel to self._history
        self._watched_blendshape: str = BLENDSHAPE_NAMES[0]  # set for real once BlendshapeWatch exists
        self._workload_state: dict = {}
        self._session_rows: list = []
        self._session_start: float = 0.0
        self._session_start_ns: int = 0
        self._frame_capture_ns: int = 0   # Updated by camera loop; used for latency
        self._last_record_time: float = 0.0
        self._session_ended = False
        self._event_log: list = []
        self._raw_facial_rows: list = []
        self._raw_vocal_rows:  list = []
        self._ws_clients: set = set()
        self._ws_loop = None
        self._ws_queue = None
        self._ws_enabled: bool = True

        # ── Participant info (collected before calibration each session) ────
        self._demographics: dict = {}
        self._demo_dialog = None

        # ── Calibration state ───────────────────────────────────────────────
        self._calibrating = False
        self._calibration_started_at = None
        self._calibration_seconds = 30
        self._calibration_pupil:  list[float] = []
        self._calibration_face = {"eye_ar": [], "blink_rate": [], "gaze_deviation": []}
        # Resting expression / Action Unit / Duchenne readings, keyed by name.
        # These are what make the facial channel personal: a face that registers
        # a permanent trace of sadness or a habitually furrowed brow scores from
        # there rather than being penalised for sitting still.
        self._calibration_expr: dict = {}
        self._calibration_aus:  dict = {}
        self._calibration_duchenne: list[float] = []
        # pitch_stability / energy_level / tremor_index are the original three vocal calibration buffers.
        # hnr_db, alpha_ratio, and jitter were added when eGeMAPS support was introduced so that
        # the VoicePanel can show "vs calibration" deltas for the two new metric boxes (HNR and Jitter).
        # spectral_flux was added when vocal scoring became baseline-relative.
        self._calibration_vocal = {"pitch_stability": [], "energy_level": [], "tremor_index": [],
                                   "hnr_db": [], "alpha_ratio": [], "jitter": [],
                                   "spectral_flux": []}
        self._calibration_hrv: list[float] = []
        self._calibration_rmssd: list[float] = []
        self._calibration_baseline: dict = {}

        # ── Behavioural flag state ──────────────────────────────────────────
        self._flag_cooldowns: dict = {}
        self._last_flag_total = None
        self._session_flags: list = []

        # ── Calibration coverage ────────────────────────────────────────────
        self._cal_frames_total = 0
        self._cal_frames_face  = 0
        self._baseline_coverage = None

        # ── Phase tracking (researcher-marked trust-protocol phases) ─────────
        # Researcher marks each transition manually (hotkeys 1/2/3 or the
        # TopStrip button) — the system never infers phase changes from the
        # score. Order is enforced: Establishment → Violation → Recovery.
        self._phase_defs = [
            ("establishment", "Trust Establishment", ACCENT),
            ("violation",      "Trust Violation",     DANGER),
            ("recovery",        "Trust Recovery",      C_WORKLOAD),
        ]
        self._phase_index = -1                  # -1 = no phase marked yet this session
        self._phase_segments: list[dict] = []    # [{key,label,color,start_s,end_s}], end_s=None while ongoing

        # ── Camera bookkeeping ──────────────────────────────────────────────
        self._available_cameras: list[int] = []
        self._camera_idx_pos = 0
        self._cam_ok = False
        # Everything known about the camera in use, from the last scan:
        # {"index", "name", "transport", "label"}. Recorded in the export so a
        # session can be traced back to the lens it was captured through.
        self._camera_info: dict = {}
        self._cam_dialog = None
        self._cam_threads_started = False
        self._mic_ok = False

        # ── Data directories ─────────────────────────────────────────────────
        self._data_dir      = Path.home() / "Desktop" / "trust-dashboard"
        self._session_dir   = self._data_dir / "session-data"   # JSON + Excel exports
        self._recordings_dir = self._data_dir / "recordings"    # video + thumbnails
        for d in (self._data_dir, self._session_dir, self._recordings_dir):
            d.mkdir(parents=True, exist_ok=True)

        # ── Recording pipeline ───────────────────────────────────────────────
        self._writer: cv2.VideoWriter | None = None
        self._writer_lock = threading.Lock()
        self._recording_path: Path | None = None
        self._session_id: str = ""

        # ── Latest scores (read by camera loop for recording overlay) ────────
        self._last_scores: dict = {}   # Written by _update_body; read by _camera_loop

        # ── Persistence ─────────────────────────────────────────────────────
        self._sessions_file = self._session_dir / "sessions.json"

        # ── Build the UI ────────────────────────────────────────────────────
        self._build_ui()
        self._show_overview()

        # ── Main UI tick (60 ms — comfortable for eye, plenty fast for data)
        self._tick = QTimer(self)
        self._tick.timeout.connect(self._update_body)
        self._tick.start(60)

    # ════════════════════════════════════════════════════════════════════════
    # UI assembly
    # ════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        self._stack = QStackedWidget()
        self.setCentralWidget(self._stack)

        # Live dashboard screen — built once, reused
        self._live = self._build_live_dashboard()
        self._stack.addWidget(self._live)   # index 0

        # The other screens (overview/calibration/summary) are constructed on
        # demand and swapped in via the stack.
        self._overview = None
        self._cal = None
        self._sum = None

    def _build_live_dashboard(self) -> QWidget:
        """Top strip + 3-column main row + history chart + footer."""
        root = QWidget()
        root.setStyleSheet(f"background: {BG};")
        rl = QVBoxLayout(root)
        rl.setContentsMargins(sp(0), sp(0), sp(0), sp(0))
        rl.setSpacing(sp(0))

        # Top strip
        self.top = TopStrip()
        self.top.end_session_clicked.connect(self._end_session)
        self.top.phase_advance_clicked.connect(self._on_phase_button_clicked)
        rl.addWidget(self.top)

        # Stage — padded content area
        stage = QWidget()
        stage.setStyleSheet(f"background: {BG};")
        sl = QVBoxLayout(stage)
        sl.setContentsMargins(sp(20), sp(18), sp(20), sp(18))
        sl.setSpacing(sp(14))

        # Row 1: camera | score | voice
        row1 = QHBoxLayout()
        row1.setSpacing(sp(14))

        self.cam_panel = CameraPanel()
        self.cam_panel.switch_camera_clicked.connect(self._switch_camera)
        self.cam_panel.setFixedWidth(sp(300))

        self.score_panel = ScorePanel()
        self.score_panel.setFixedWidth(sp(520))

        self.voice_panel = VoicePanel()
        self.voice_panel.setMaximumWidth(sp(380))

        # Cap height so panels don't over-stretch on tall windows
        for _p in (self.cam_panel, self.score_panel, self.voice_panel):
            _p.setMaximumHeight(sp(600))

        row1.addWidget(self.cam_panel)
        row1.addWidget(self.score_panel)
        row1.addWidget(self.voice_panel, 1)
        sl.addLayout(row1, 0)

        # Row 1.5: blendshape watch — full-width, short, sits between the
        # summary panels and the long-form history chart below.
        self.blendshape_watch = BlendshapeWatch(BLENDSHAPE_NAMES)
        self.blendshape_watch.blendshape_changed.connect(self._on_blendshape_changed)
        self._watched_blendshape = self.blendshape_watch.current_blendshape()
        sl.addWidget(self.blendshape_watch, 0)

        # Row 2: history chart full-width
        self.history_chart = HistoryChart()
        self.history_chart.setMinimumHeight(sp(220))
        sl.addWidget(self.history_chart, 1)

        # Wrap stage + flag sidebar in a middle HBox
        self.flag_sidebar = FlagSidebar()
        middle = QHBoxLayout()
        middle.setContentsMargins(sp(0), sp(0), sp(0), sp(0))
        middle.setSpacing(sp(0))
        middle.addWidget(stage, 1)
        middle.addWidget(self.flag_sidebar)
        rl.addLayout(middle, 1)

        # Footer
        rl.addWidget(Footer())
        return root

    # ════════════════════════════════════════════════════════════════════════
    # Screen routing
    # ════════════════════════════════════════════════════════════════════════
    def _show_overview(self):
        """Build (or rebuild) and show the overview/landing page."""
        if self._overview is not None:
            self._stack.removeWidget(self._overview)
            self._overview.deleteLater()
        self._overview = OverviewScreen(self._sessions_file)
        self._overview.start_clicked.connect(self._start_session)
        self._overview.session_clicked.connect(self._show_past_session_summary)
        self._stack.addWidget(self._overview)
        self._stack.setCurrentWidget(self._overview)

    def _show_live(self):
        self._stack.setCurrentWidget(self._live)

    def _show_calibration(self):
        if self._cal is not None:
            self._stack.removeWidget(self._cal)
            self._cal.deleteLater()
        self._cal = CalibrationOverlay()
        self._cal.start_clicked.connect(self._begin_calibration)
        self._cal.skip_clicked.connect(self._finish_calibration_now)
        self._stack.addWidget(self._cal)
        self._stack.setCurrentWidget(self._cal)

    def _show_summary(self, stats: dict, export_handler=None):
        """export_handler lets callers repoint the Export button — the
        just-ended live session exports its in-memory rows, but a summary
        opened from the overview's session list has no in-memory rows to
        export, so it reveals the already auto-saved Excel file instead."""
        if self._sum is not None:
            self._stack.removeWidget(self._sum)
            self._sum.deleteLater()
        self._sum = SessionSummary()
        self._sum.populate(stats)
        self._sum.back_clicked.connect(self._back_to_overview)
        self._sum.export_clicked.connect(export_handler or self._export_csv)
        self._stack.addWidget(self._sum)
        self._stack.setCurrentWidget(self._sum)

    def _show_past_session_summary(self, sess: dict):
        """Open the read-only summary for a previously recorded session,
        selected by clicking a card in the overview's session list."""
        self._show_summary(sess, export_handler=lambda: self._reveal_session_export(sess))

    def _reveal_session_export(self, sess: dict):
        """Open the Excel file that was auto-saved for this session at the
        time it ended, rather than trying to rebuild one from (empty)
        live session data."""
        session_id = sess.get("session_id", "")
        path = self._session_dir / f"trust-session-{session_id}.xlsx"
        if path.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))
        else:
            QMessageBox.information(
                self, "Export not found",
                f"No saved Excel export was found for this session:\n{path}",
            )

    # ════════════════════════════════════════════════════════════════════════
    # Session lifecycle
    # ════════════════════════════════════════════════════════════════════════
    def _start_session(self):
        """User clicked Start on the overview — collect participant info
        (sex, age, culture) before anything else, then open calibration."""
        self._demo_dialog = DemographicsDialog(self)
        self._demo_dialog.completed.connect(self._on_demographics_complete)
        self._demo_dialog.open()

    def _on_demographics_complete(self, result: dict | None):
        """Fired when the participant-info dialog finishes. `result` is a
        {"sex", "age", "culture"} dict, or None if the researcher cancelled —
        in which case we stay on the overview and never touch the camera."""
        self._demo_dialog = None
        if result is None:
            return
        self._demographics = result
        # Confirm the camera before calibration rather than after: a baseline
        # recorded through the wrong lens is a wasted 30 seconds and, worse, a
        # wrong baseline for the whole session.
        self._open_camera_picker(
            on_chosen=lambda info: self._begin_session_setup(info["index"]),
            on_cancelled=self._back_to_overview,
        )

    def _begin_session_setup(self, camera_index: int):
        """Camera confirmed — open calibration and bring the sensors up."""
        self._show_calibration()
        # Start camera + audio in background so the calibration preview
        # already has a live feed when the user clicks Start Calibration.
        self._start_camera(camera_index)
        self._start_audio()
        self.hrv.start()  # scans/connects to a BLE HR strap (e.g. Polar H10) in the background

    def _begin_calibration(self):
        """User clicked Start Calibration inside the overlay."""
        # Clear every buffer first. These persist on the window, so without this
        # a second session in the same run would average its baseline together
        # with the previous participant's readings.
        self._calibration_pupil = []
        self._calibration_face = {k: [] for k in self._calibration_face}
        self._calibration_vocal = {k: [] for k in self._calibration_vocal}
        self._calibration_expr = {}
        self._calibration_aus = {}
        self._calibration_duchenne = []
        self._calibration_hrv = []
        self._calibration_rmssd = []

        self._calibration_started_at = time.time()
        self._calibrating = True
        self._session_ended = False
        self._cal_frames_total = 0
        self._cal_frames_face  = 0

    def _finish_calibration_now(self):
        """User clicked Skip."""
        self._calibrating = False
        self._calibration_started_at = None
        self._enter_live_session()

    def _collect_calibration_samples(self, face_data, vocal_data):
        if face_data and face_data.get("detected"):
            self._calibration_face["eye_ar"].append(float(face_data.get("eye_ar", 0.27)))
            self._calibration_face["blink_rate"].append(float(face_data.get("blink_rate", 15.0)))
            self._calibration_face["gaze_deviation"].append(float(face_data.get("gaze_deviation", 0.0)))
            p = face_data.get("pupil_norm")
            if p is not None:
                self._calibration_pupil.append(float(p))
            # The resting face itself: every emotion and Action Unit MediaPipe
            # reports while the person sits relaxed. Collected per key so a
            # partial frame never drops a signal that other frames did have.
            for key, val in (face_data.get("expressions") or {}).items():
                self._calibration_expr.setdefault(key, []).append(float(val))
            for key, val in (face_data.get("aus") or {}).items():
                self._calibration_aus.setdefault(key, []).append(float(val))
            self._calibration_duchenne.append(float(face_data.get("duchenne", 0.0)))
        # Voice samples are only meaningful while the person is actually
        # speaking. Averaging in silent frames would drag the resting loudness
        # toward zero and make every later utterance look like shouting.
        if vocal_data and vocal_data.get("is_speaking"):
            self._calibration_vocal["pitch_stability"].append(float(vocal_data.get("pitch_stability", 0.5)))
            self._calibration_vocal["energy_level"].append(float(vocal_data.get("energy_level", 0.0)))
            self._calibration_vocal["tremor_index"].append(float(vocal_data.get("tremor_index", 0.0)))
            # Only append the eGeMAPS features when they are non-zero; 0.0 is the sentinel
            # returned by the legacy fallback path (opensmile not installed) and must not
            # pollute the calibration baseline with fake zeros.
            if vocal_data.get("hnr_db", 0.0) != 0.0:
                self._calibration_vocal["hnr_db"].append(float(vocal_data["hnr_db"]))
            if vocal_data.get("alpha_ratio", 0.0) != 0.0:
                self._calibration_vocal["alpha_ratio"].append(float(vocal_data["alpha_ratio"]))
            if vocal_data.get("jitter", 0.0) != 0.0:
                self._calibration_vocal["jitter"].append(float(vocal_data["jitter"]))
            if vocal_data.get("spectral_flux", 0.0) > 0.0:
                self._calibration_vocal["spectral_flux"].append(float(vocal_data["spectral_flux"]))
        # Only count a beat once the analyzer has a real RMSSD reading — while
        # "connected" but still filling its rolling R-R window it reports the
        # stub score, which must not pollute the HRV baseline.
        hrv_display = self.hrv.get_display()
        if hrv_display.get("rmssd_ms") is not None:
            self._calibration_hrv.append(float(hrv_display["score"]))
            self._calibration_rmssd.append(float(hrv_display["rmssd_ms"]))

    @staticmethod
    def _mean_or(values, fallback):
        return sum(values) / len(values) if values else fallback

    def _enter_live_session(self):
        """Calibration complete (or skipped) — switch to live dashboard."""
        self._calibration_baseline = {
            "face_eye_ar":            self._mean_or(self._calibration_face["eye_ar"], 0.27),
            "face_blink_rate":        self._mean_or(self._calibration_face["blink_rate"], 15.0),
            "face_gaze_deviation":    self._mean_or(self._calibration_face["gaze_deviation"], 0.0),
            "voice_pitch_stability":  self._mean_or(self._calibration_vocal["pitch_stability"], 0.5),
            "voice_energy_level":     self._mean_or(self._calibration_vocal["energy_level"], 0.0),
            "voice_tremor_index":     self._mean_or(self._calibration_vocal["tremor_index"], 0.0),
            # None as the fallback means "no eGeMAPS data collected" (opensmile not installed);
            # VoicePanel.update_metrics checks for None before showing deltas on the new HNR/Jitter boxes.
            "voice_hnr_db":           self._mean_or(self._calibration_vocal["hnr_db"], None),
            "voice_alpha_ratio":      self._mean_or(self._calibration_vocal["alpha_ratio"], None),
            "voice_jitter":           self._mean_or(self._calibration_vocal["jitter"], None),
            # None means no RMSSD-backed reading arrived during the window (no strap worn/
            # connected) — same "no data" convention as the eGeMAPS voice fields above.
            "hrv_score":              self._mean_or(self._calibration_hrv, None),
            "hrv_rmssd_ms":           self._mean_or(self._calibration_rmssd, None),
        }
        # ── Hand the measured resting readings to the analysers ─────────────
        #
        # Everything below uses None to mean "this signal was never captured"
        # (no face detected, nobody spoke, no strap worn), which leaves that one
        # signal on its population default while the rest still personalise.
        #
        # RMSSD goes to the HRV analyser rather than the trust engine, because
        # that channel is scored from R-R intervals inside the analyser.
        resting_rmssd = self._mean_or(self._calibration_rmssd, None)
        self.hrv.set_baseline_rmssd(resting_rmssd)

        measured = {
            "eye_ar":          self._mean_or(self._calibration_face["eye_ar"], None),
            "blink_rate":      self._mean_or(self._calibration_face["blink_rate"], None),
            "gaze_deviation":  self._mean_or(self._calibration_face["gaze_deviation"], None),
            "pitch_stability": self._mean_or(self._calibration_vocal["pitch_stability"], None),
            "energy_level":    self._mean_or(self._calibration_vocal["energy_level"], None),
            "tremor_index":    self._mean_or(self._calibration_vocal["tremor_index"], None),
            "alpha_ratio":     self._mean_or(self._calibration_vocal["alpha_ratio"], None),
            "spectral_flux":   self._mean_or(self._calibration_vocal["spectral_flux"], None),
            "duchenne":        self._mean_or(self._calibration_duchenne, None),
            "expressions":     {k: self._mean_or(v, None)
                                for k, v in self._calibration_expr.items()},
            "aus":             {k: self._mean_or(v, None)
                                for k, v in self._calibration_aus.items()},
        }
        # Resting HRV score: 50 once a real RMSSD baseline exists, because the
        # analyser now centres this user's resting RMSSD there. Without one the
        # analyser reports its constant stub, so that constant *is* the resting
        # value and subtracting it keeps the channel neutral instead of letting
        # a fixed 65 quietly lift every total by a quarter of the difference.
        hrv_resting = 50.0 if resting_rmssd else float(self.hrv.get_score())

        self.trust = TrustEngine()
        self.trust.apply_calibration(measured, hrv_resting=hrv_resting)
        self._history = {k: [] for k in ("total", "facial", "vocal", "gaze", "hrv")}
        self._history_t = []
        self._phase_index = -1
        self._phase_segments = []
        self.top.reset_phase()
        if self._calibration_pupil:
            self.workload.set_baseline(sum(self._calibration_pupil) / len(self._calibration_pupil))
        self._calibrating = False
        self._session_id = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        self._best_thumb_frame = None
        self._best_thumb_conf = -1.0
        self._session_start = time.time()
        self._session_start_ns = time.time_ns()
        self._session_rows = []
        self._raw_facial_rows = []
        self._raw_vocal_rows  = []
        self._event_log = []
        self.log_event("sync", "session_start")
        self._show_sync_flash("SESSION START")
        self._start_recording()

        # Write session start sidecar for WorldCam/external sync
        _sidecar = self._recordings_dir / f"{self._session_id}_start.json"
        try:
            import json as _json
            _sidecar.write_text(_json.dumps({
                "session_id": self._session_id,
                "start_ns": self._session_start_ns,
                "start_wall": datetime.now().isoformat(),
                "score_version": SCORE_VERSION,
                "score_config": SCORE_CONFIG,
                "active_channels": [ch for ch, a in self.trust._active.items() if a],
            }, indent=2))
        except Exception:
            pass

        # Mark the gauge with this user's resting score. Pushing the engine's own
        # resting sample back through the live engine measures it the same way
        # every live frame is measured, rather than asserting a number — it lands
        # on 50 whenever calibration worked, so a marker sitting anywhere else is
        # a visible sign that something in the window did not take.
        probe_engine = TrustEngine()
        probe_engine.input_baseline = self.trust.input_baseline
        probe_engine.baseline = dict(self.trust.baseline)
        face_probe, vocal_probe = self.trust.resting_samples()
        for _ in range(20):
            probe_engine.update(face_probe, vocal_probe, hrv_resting)
        result = probe_engine.update(face_probe, vocal_probe, hrv_resting)
        self._baseline_total = int(result["total"])
        self.score_panel.gauge.setBaseline(self._baseline_total)

        if self._cal_frames_total > 0:
            self._baseline_coverage = 100.0 * self._cal_frames_face / self._cal_frames_total
            self.score_panel.set_baseline_quality(self._baseline_coverage)
        else:
            self._baseline_coverage = None
            self.score_panel.hide_baseline_quality()
        self.flag_sidebar.clear_flags()
        self._flag_cooldowns = {}
        self._last_flag_total = None
        self._session_flags = []

        self._show_live()
        self._start_ws_server()

    def _end_session(self):
        if len(self._session_rows) < 2:
            QMessageBox.information(
                self, "Not enough data",
                "Not enough data yet — wait a few seconds before ending the session.",
            )
            return
        self.log_event("sync", "session_end")
        self._show_sync_flash("SESSION END")
        self._session_ended = True
        # Stop recording before computing stats so the writer is flushed
        rec_path, thumb_path = self._stop_recording()
        stats = self._compute_session_stats()
        def _rel(p):
            """Store recording paths relative to data_dir for portability."""
            if not p:
                return None
            try:
                return str(Path(p).relative_to(self._data_dir))
            except Exception:
                return str(p)

        stats["recording_path"] = _rel(rec_path)
        stats["thumbnail_path"] = _rel(thumb_path)
        stats["session_id"]     = self._session_id
        self._save_session(stats)

        # Auto-save the session Excel export to the session directory.
        auto_excel = self._session_dir / f"trust-session-{self._session_id}.xlsx"
        try:
            self._build_excel(str(auto_excel))
            print(f"[export] Auto-saved Excel → {auto_excel}", flush=True)
        except Exception as e:
            print(f"[export] Auto-save failed: {e}", flush=True)

        # Transcode the recording to H.264 in the background so it plays back
        # cleanly in QuickTime. Runs unconditionally, independent of the UI.
        if rec_path and rec_path.exists():
            self._transcode_recording(rec_path, getattr(self, "_rec_actual_fps", None))

        summary_stats = dict(stats)
        summary_stats["recording_path"] = str(rec_path) if rec_path else None
        summary_stats["thumbnail_path"] = str(thumb_path) if thumb_path else None
        self._show_summary(summary_stats)

    def _back_to_overview(self):
        self._session_ended = False
        self._session_rows = []
        self._history = {k: [] for k in ("total", "facial", "vocal", "gaze", "hrv")}
        self._history_t = []
        self.flag_sidebar.clear_flags()
        self._flag_cooldowns = {}
        self._last_flag_total = None
        self._session_flags = []
        self.score_panel.hide_baseline_quality()
        self._phase_index = -1
        self._phase_segments = []
        self._show_overview()

    # ════════════════════════════════════════════════════════════════════════
    # Background video transcode
    # ════════════════════════════════════════════════════════════════════════
    def _transcode_recording(self, video_path: Path, actual_fps: float | None = None):
        """
        Re-encode the just-recorded .mp4 (written with the mp4v fourcc for
        reliable capture) to H.264 in a daemon background thread, so it plays
        back cleanly in QuickTime and other players. Runs unconditionally
        after every session, independent of any analysis pipeline; the UI is
        never blocked waiting for it.

        *actual_fps* is the true frames-written / real-elapsed-time rate
        measured during recording (see `_stop_recording`), which can differ
        from the container's nominal fps if capture throughput dipped during
        the session. Passing it as an input `-r` override re-times every
        frame at the real rate so the output duration matches the real
        session length instead of playing back sped up.
        """
        def _worker(path: Path):
            tmp = path.with_suffix(".h264_tmp.mp4")
            try:
                cmd = ["ffmpeg", "-y"]
                if actual_fps and actual_fps > 0:
                    cmd += ["-r", f"{actual_fps:.3f}"]
                cmd += ["-i", str(path),
                        "-c:v", "libx264", "-preset", "fast", "-crf", "23",
                        "-movflags", "+faststart", "-an", str(tmp)]
                r = subprocess.run(cmd, capture_output=True, timeout=600)
                if r.returncode == 0 and tmp.exists() and tmp.stat().st_size > 0:
                    tmp.replace(path)
                    print("[rec] Transcoded to H.264 for QuickTime compatibility.", flush=True)
                else:
                    tmp.unlink(missing_ok=True)
                    print("[rec] H.264 transcode failed — keeping original.", flush=True)
            except Exception as e:
                tmp.unlink(missing_ok=True)
                print(f"[rec] H.264 transcode error: {e}", flush=True)

        threading.Thread(target=_worker, args=(video_path,), daemon=True).start()
        print("[rec] Background H.264 transcode started.", flush=True)

    # ════════════════════════════════════════════════════════════════════════
    # Recording helpers
    # ════════════════════════════════════════════════════════════════════════
    def _start_recording(self):
        """Open a VideoWriter for the current session. Called from the main
        thread after calibration; must run while self._cap is already open."""
        with self._cap_lock:
            cap = self._cap
            if cap is None or not cap.isOpened():
                print("[rec] No camera — recording disabled")
                return
            w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self._recordings_dir.mkdir(exist_ok=True)
        path = self._recordings_dir / f"{self._session_id}.mp4"
        if w < 1 or h < 1:
            w, h = 1280, 720
        # Use the empirically-measured delivery rate (set in _start_camera /
        # _switch_camera) rather than cap.get(CAP_PROP_FPS) — the driver
        # can report a fps that doesn't match what's actually delivered,
        # and stamping the container with the wrong fps is what made
        # recordings play back sped up.
        fps = self._measured_cam_fps
        if fps < 10 or fps > 120:
            fps = 30.0
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")   # reliable write; transcoded to H.264 post-session
        writer = cv2.VideoWriter(str(path), fourcc, fps, (w, h))
        if writer.isOpened():
            with self._writer_lock:
                self._writer = writer
                self._rec_frame_count = 0
                self._rec_start_ns = time.time_ns()
            self._recording_path = path
            print(f"[rec] Recording → {path}  ({w}×{h} @ {fps:.1f}fps)")
        else:
            print("[rec] VideoWriter failed to open — no recording")

    def _stop_recording(self) -> tuple["Path | None", "Path | None"]:
        """Release the writer and save the best-thumbnail JPEG.
        Returns (recording_path, thumbnail_path), both may be None."""
        # ── 1. Grab and clear the writer atomically ──────────────────────────
        with self._writer_lock:
            writer = self._writer
            self._writer = None
            frame_count = getattr(self, "_rec_frame_count", 0)
            start_ns = getattr(self, "_rec_start_ns", None)
        stop_ns = time.time_ns()
        # The container was stamped with a nominal fps measured before the
        # session started; actual throughput during recording (contended by
        # the analysis loop, overlay drawing, etc.) can run slower, which is
        # what made played-back recordings finish sooner than the real
        # session length. Recompute the true fps from frames actually
        # written over the real elapsed time, and use that to correct the
        # container's frame rate at transcode time.
        self._rec_actual_fps = self._measured_cam_fps
        if start_ns is not None and frame_count > 1:
            elapsed_s = (stop_ns - start_ns) / 1e9
            if elapsed_s > 0:
                self._rec_actual_fps = max(1.0, min(frame_count / elapsed_s, 60.0))
        if writer is not None:
            try:
                writer.release()
            except Exception as e:
                print(f"[rec] writer.release() error: {e}")

        rec_path = self._recording_path
        self._recording_path = None

        # ── 2. Extract first frame as thumbnail using ffmpeg ────────────────
        if not rec_path or not self._session_id:
            return rec_path, None

        self._recordings_dir.mkdir(exist_ok=True)
        thumb_path = self._recordings_dir / f"{self._session_id}.jpg"
        try:
            subprocess.run(
                ["ffmpeg", "-y", "-i", str(rec_path),
                 "-vframes", "1",
                 "-vf", "scale=640:360:force_original_aspect_ratio=decrease,"
                        "pad=640:360:(ow-iw)/2:(oh-ih)/2",
                 "-q:v", "2",
                 str(thumb_path)],
                capture_output=True, timeout=30,
            )
            print(f"[rec] Thumbnail → {thumb_path}")
            return rec_path, thumb_path
        except Exception as e:
            print(f"[rec] Thumbnail save failed: {e}")
            return rec_path, None

    # ════════════════════════════════════════════════════════════════════════
    # Recording overlay
    # ════════════════════════════════════════════════════════════════════════

    # MediaPipe face mesh connection sets — hardcoded from the MediaPipe topology.
    # mediapipe.solutions was removed in MediaPipe 0.10+, so we define the
    # static index pairs directly.  The 478-point topology is fixed and will
    # not change between MediaPipe versions.

    # Face oval — 36 edges, clockwise from top-centre
    _FACE_OVAL: frozenset = frozenset([
        (10, 338), (338, 297), (297, 332), (332, 284), (284, 251), (251, 389),
        (389, 356), (356, 454), (454, 323), (323, 361), (361, 288), (288, 397),
        (397, 365), (365, 379), (379, 378), (378, 400), (400, 377), (377, 152),
        (152, 148), (148, 176), (176, 149), (149, 150), (150, 136), (136, 172),
        (172,  58), ( 58, 132), (132,  93), ( 93, 234), (234, 127), (127, 162),
        (162,  21), ( 21,  54), ( 54, 103), (103,  67), ( 67, 109), (109,  10),
    ])

    # Left eye outline (16 edges, MediaPipe indices for the subject's left eye)
    _MESH_L_EYE: frozenset = frozenset([
        (362, 382), (382, 381), (381, 380), (380, 374), (374, 373), (373, 390),
        (390, 249), (249, 263), (263, 466), (466, 388), (388, 387), (387, 386),
        (386, 385), (385, 384), (384, 398), (398, 362),
    ])

    # Right eye outline (16 edges)
    _MESH_R_EYE: frozenset = frozenset([
        ( 33,   7), (  7, 163), (163, 144), (144, 145), (145, 153), (153, 154),
        (154, 155), (155, 133), (133, 173), (173, 157), (157, 158), (158, 159),
        (159, 160), (160, 161), (161, 246), (246,  33),
    ])

    # Iris circles — indices 468-471 (left) and 472-475 (right)
    _MESH_L_IRIS: frozenset = frozenset([(468, 469), (469, 470), (470, 471), (471, 468)])
    _MESH_R_IRIS: frozenset = frozenset([(472, 473), (473, 474), (474, 475), (475, 472)])

    # Lips — outer contour
    _LIPS_OUTER: frozenset = frozenset([
        ( 61, 185), (185,  40), ( 40,  39), ( 39,  37), ( 37,   0), (  0, 267),
        (267, 269), (269, 270), (270, 409), (409, 291), (291, 375), (375, 321),
        (321, 405), (405, 314), (314,  17), ( 17,  84), ( 84, 181), (181,  91),
        ( 91, 146), (146,  61),
    ])

    @staticmethod
    def _draw_connections(frame, pts, connections, colour, thickness=1):
        """Draw a set of (idx_a, idx_b) mesh connections onto frame."""
        if connections is None:
            return
        for a, b in connections:
            if a < len(pts) and b < len(pts):
                cv2.line(frame, pts[a], pts[b], colour, thickness, cv2.LINE_AA)

    def _draw_face_mesh(self, frame: np.ndarray, face_data: dict | None) -> None:
        """Draw the face mesh landmarks and feature outlines onto *frame* in-place."""
        lms_norm = (face_data or {}).get("landmarks_norm")
        if not lms_norm:
            return
        h, w = frame.shape[:2]
        pts = [(int(x * w), int(y * h)) for x, y in lms_norm]

        dot_overlay = frame.copy()
        for px_, py_ in pts[:468]:
            cv2.circle(dot_overlay, (px_, py_), 1, (255, 255, 255), -1, cv2.LINE_AA)
        cv2.addWeighted(dot_overlay, 0.45, frame, 0.55, 0, frame)

        self._draw_connections(frame, pts, self._FACE_OVAL,   (255, 255,   0), thickness=1)
        self._draw_connections(frame, pts, self._LIPS_OUTER,  ( 60, 100, 255), thickness=1)
        self._draw_connections(frame, pts, self._MESH_L_EYE,  (255,  80,   0), thickness=2)
        self._draw_connections(frame, pts, self._MESH_R_EYE,  (255,   0, 230), thickness=2)
        self._draw_connections(frame, pts, self._MESH_L_IRIS, (255,  80,   0), thickness=2)
        self._draw_connections(frame, pts, self._MESH_R_IRIS, (255,   0, 230), thickness=2)

    def _draw_recording_overlay(self, frame: np.ndarray, face_data: dict | None,
                                 scores: dict | None) -> np.ndarray:
        """Draw the blendshape emotion panel onto a recording frame."""
        h, w = frame.shape[:2]

        # ── Emotion definitions: label, BGR colour ────────────────────────────
        EMOTIONS = [
            ("happy",     (86,  211,  86)),   # Green
            ("neutral",   (180, 180, 180)),   # Grey
            ("surprised", ( 50, 210, 210)),   # Yellow
            ("sad",       (180, 130,  80)),   # Steel blue
            ("angry",     ( 60,  60, 220)),   # Red
            ("fearful",   ( 60, 150, 220)),   # Orange
            ("disgusted", (160,  60, 160)),   # Purple
            ("contempt",  ( 50,  50, 170)),   # Dark red
        ]

        # ── Panel geometry ────────────────────────────────────────────────────
        PAD        = 10    # Inner padding
        ROW_H      = 26    # Height of each emotion row
        BAR_MAX_W  = 110   # Maximum bar width in pixels
        LABEL_W    = 72    # Width reserved for the emotion label
        VAL_W      = 34    # Width reserved for the numeric value
        PANEL_W    = PAD + LABEL_W + 6 + BAR_MAX_W + 6 + VAL_W + PAD   # ~248 px
        HEADER_H   = 38    # Title row height
        FOOTER_H   = 34    # Trust score row height
        PANEL_H    = PAD + HEADER_H + len(EMOTIONS) * ROW_H + FOOTER_H + PAD

        px = w - PANEL_W - 16   # Right-align with a small margin
        py = 16                  # Top margin

        # ── Semi-transparent dark background ──────────────────────────────────
        overlay = frame.copy()
        cv2.rectangle(overlay, (px, py), (px + PANEL_W, py + PANEL_H),
                      (20, 20, 20), cv2.FILLED)
        cv2.addWeighted(overlay, 0.72, frame, 0.28, 0, frame)

        # ── Thin border ───────────────────────────────────────────────────────
        cv2.rectangle(frame, (px, py), (px + PANEL_W, py + PANEL_H),
                      (80, 80, 80), 1)

        # ── Header: "BLENDSHAPE EMOTIONS" ─────────────────────────────────────
        cv2.putText(frame, "BLENDSHAPE EMOTIONS",
                    (px + PAD, py + PAD + 14),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.42, (200, 200, 200), 1, cv2.LINE_AA)
        # Thin separator line under header
        sep_y = py + PAD + HEADER_H - 4
        cv2.line(frame, (px + PAD, sep_y), (px + PANEL_W - PAD, sep_y), (70, 70, 70), 1)

        # ── Emotion bars ──────────────────────────────────────────────────────
        expressions = (face_data or {}).get("expressions", {})
        dominant    = (face_data or {}).get("dominant", "")

        bar_x  = px + PAD + LABEL_W + 6   # Left edge of the bar area
        val_x  = bar_x + BAR_MAX_W + 4    # Left edge of the value column

        for i, (emotion, colour) in enumerate(EMOTIONS):
            row_y = py + PAD + HEADER_H + i * ROW_H
            score = float(expressions.get(emotion, 0.0))

            # Label — bold-ish by drawing twice for weight
            label_colour = colour if emotion == dominant else (160, 160, 160)
            cv2.putText(frame, emotion,
                        (px + PAD, row_y + 17),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, label_colour, 1, cv2.LINE_AA)

            # Grey track
            cv2.rectangle(frame,
                          (bar_x, row_y + 6),
                          (bar_x + BAR_MAX_W, row_y + 19),
                          (55, 55, 55), cv2.FILLED)

            # Coloured fill — width proportional to score
            fill_w = int(BAR_MAX_W * score)
            if fill_w > 0:
                cv2.rectangle(frame,
                              (bar_x, row_y + 6),
                              (bar_x + fill_w, row_y + 19),
                              colour, cv2.FILLED)

            # Highlight bar for dominant emotion
            if emotion == dominant:
                cv2.rectangle(frame,
                              (bar_x, row_y + 6),
                              (bar_x + BAR_MAX_W, row_y + 19),
                              colour, 1)

            # Numeric value
            cv2.putText(frame, f"{score:.2f}",
                        (val_x, row_y + 17),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.36, (180, 180, 180), 1, cv2.LINE_AA)

        # ── Footer: trust score ───────────────────────────────────────────────
        foot_y = py + PAD + HEADER_H + len(EMOTIONS) * ROW_H + 6
        cv2.line(frame, (px + PAD, foot_y), (px + PANEL_W - PAD, foot_y), (70, 70, 70), 1)

        total = int((scores or {}).get("total", 0))
        # Pick a colour that matches the trust_label bands
        if   total >= 82: tc = ( 80, 222, 74)    # Green
        elif total >= 64: tc = ( 57, 211, 52)     # Teal
        elif total >= 46: tc = (250, 165, 96)     # Blue
        elif total >= 28: tc = ( 50, 147, 251)    # Orange
        else:             tc = (113, 129, 248)    # Red

        cv2.putText(frame, f"TRUST  {total}",
                    (px + PAD, foot_y + 22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.50, tc, 1, cv2.LINE_AA)

        # Dominant emotion label next to trust score
        if dominant:
            cv2.putText(frame, dominant.upper(),
                        (px + PAD + 100, foot_y + 22),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.36, (140, 140, 140), 1, cv2.LINE_AA)

        return frame

    # ════════════════════════════════════════════════════════════════════════
    # Camera + analysis threads
    # ════════════════════════════════════════════════════════════════════════
    # ── Camera preference persistence ────────────────────────────────────────
    # Saves the last-used camera index to a small JSON file so the same camera
    # is selected automatically on the next launch.

    def _load_camera_pref(self) -> int | None:
        """Return the saved preferred camera index, or None if not set."""
        try:
            p = self._data_dir / "camera_pref.json"
            if p.exists():
                return int(json.loads(p.read_text()).get("index", -1)) or None
        except Exception:
            pass
        return None

    def _save_camera_pref(self, index: int):
        """Persist the chosen camera index for next launch."""
        try:
            p = self._data_dir / "camera_pref.json"
            p.write_text(json.dumps({"index": index}))
        except Exception:
            pass

    @staticmethod
    def _warm_up_camera_permission():
        """Open index 0 once so macOS raises its camera-permission prompt.

        Without this the first real open happens inside the picker's worker
        thread, where the prompt is easy to miss and every probe fails until it
        is answered. Harmless no-op on Windows and Linux. stderr is muted
        because a denied open prints a wall of driver noise.

        Note the settle pause at the end: this warm-up opens and releases
        index 0, and the scan that follows immediately reopens it. AVFoundation
        does not hand the device straight back, so without the pause the
        machine's built-in camera reports itself as unopenable — the one device
        the researcher is most likely to want.
        """
        old_err = None
        try:
            old_err = os.dup(2)
            os.dup2(os.open(os.devnull, os.O_WRONLY), 2)
        except Exception:
            old_err = None
        try:
            cap = cv2.VideoCapture(0, _CAM_BACKEND)
            cap.release()
        except Exception:
            pass
        finally:
            if old_err is not None:
                try:
                    os.dup2(old_err, 2); os.close(old_err)
                except Exception:
                    pass
        time.sleep(0.4)

    def _scan_cameras(self) -> list[dict]:
        """Full device scan; also refreshes _available_cameras."""
        self._warm_up_camera_permission()
        cameras = camera_scanner.scan_cameras(_CAM_BACKEND)
        self._available_cameras = [c["index"] for c in cameras] or [0]
        return cameras

    def _pick_camera(self) -> int:
        """Choose a camera without asking — used only as a fallback.

        The picker dialog is the normal path. This runs when the dialog is
        bypassed (a remembered camera that is still attached, or a scan that
        found nothing), and prefers the remembered index, then the first
        camera the scanner ranked, which is built-in before wired before
        wireless.
        """
        cameras = self._scan_cameras()

        saved = self._load_camera_pref()
        if saved is not None and any(c["index"] == saved for c in cameras):
            chosen = next(c for c in cameras if c["index"] == saved)
            print(f"[camera] Restored preferred index {saved} ({chosen['name']!r})")
            self._camera_info = chosen
            return saved

        if cameras:
            chosen = cameras[0]
            print(f"[camera] Selected index {chosen['index']} "
                  f"({chosen['name']!r}) — {chosen['label']}")
            self._camera_info = chosen
            return chosen["index"]

        print("[camera] No camera found — falling back to index 0")
        self._camera_info = {}
        return 0

    def _measure_cam_fps(self, n_frames: int = 20) -> float:
        """Actually clock how fast self._cap delivers frames, rather than
        trusting cap.get(CAP_PROP_FPS) — some AVFoundation drivers just echo
        back whatever fps was requested via cap.set() without truly honoring
        it (e.g. still deliver 30fps at 720p even after a 60fps request).
        Stamping the recorded VideoWriter with a nominal fps that doesn't
        match the true delivery rate is exactly what caused recordings to
        play back sped up, so we measure directly instead."""
        n_ok = 0
        t0 = time.time()
        for _ in range(n_frames):
            with self._cap_lock:
                cap = self._cap
                ok, _ = cap.read() if cap is not None else (False, None)
            if ok:
                n_ok += 1
        elapsed = time.time() - t0
        measured = (n_ok / elapsed) if elapsed > 0 and n_ok > 0 else 30.0
        return max(10.0, min(measured, 60.0))

    def _open_capture(self, index: int):
        """Point the capture at *index*, replacing whatever was open.

        Held under _cap_lock so the camera loop can never read from a device
        that is being released underneath it.
        """
        with self._cap_lock:
            if self._cap is not None:
                try:
                    self._cap.release()
                except Exception:
                    pass
            cap = cv2.VideoCapture(index, _CAM_BACKEND)
            cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
            cap.set(cv2.CAP_PROP_FPS, 60)
            self._cap = cap
        self._cam_ok = False
        self._measured_cam_fps = self._measure_cam_fps()

    def _release_capture(self):
        """Free the camera so another process — the picker's preview — can open it."""
        with self._cap_lock:
            if self._cap is not None:
                try:
                    self._cap.release()
                except Exception:
                    pass
                self._cap = None
        self._cam_ok = False

    def _apply_camera_choice(self, info: dict):
        """Open the chosen camera and reflect it in the UI + saved preference."""
        idx = int(info.get("index", 0))
        self._camera_info = dict(info)
        self._open_capture(idx)
        print(f"[cam] using index {idx} ({info.get('name', '?')} — "
              f"{info.get('label', '?')}), measured "
              f"{self._measured_cam_fps:.1f}fps actual delivery")
        self.cam_panel.set_camera_info(idx, len(self._available_cameras),
                                       info.get("name", ""), info.get("label", ""))
        self._save_camera_pref(idx)

    def _start_camera(self, index: int | None = None):
        """Start the capture and its two worker threads.

        *index* comes from the picker, which has usually opened the device
        already — in that case this only starts the threads. Without one it
        falls back to _pick_camera(), the path taken when the picker is skipped.

        Guarded on the threads rather than on the capture handle: the handle is
        legitimately None while the picker holds the device, and starting a
        second pair of loops would double every frame.
        """
        if self._cam_threads_started:
            return
        if self._cap is None:
            if index is None:
                index = self._pick_camera()
            info = self._camera_info or {"index": index, "name": f"Camera {index}",
                                         "transport": camera_scanner.UNKNOWN,
                                         "label": ""}
            self._apply_camera_choice({**info, "index": index})
        self._cam_threads_started = True
        threading.Thread(target=self._camera_loop,   daemon=True).start()
        threading.Thread(target=self._analysis_loop, daemon=True).start()

    def _open_camera_picker(self, on_chosen=None, on_cancelled=None):
        """Show the camera picker.

        The dialog previews devices by opening them itself, and on Windows a
        DirectShow device cannot be opened twice — so the live capture is
        released first and only reopened once a choice comes back. The camera
        loop tolerates the gap.
        """
        was_running = self._cap is not None
        self._release_capture()

        def _done(result):
            self._cam_dialog = None
            if result is None:
                # Cancelled: put back whatever was running before.
                if was_running:
                    self._apply_camera_choice(
                        self._camera_info or {"index": self._load_camera_pref() or 0})
                if on_cancelled is not None:
                    on_cancelled()
                return
            self._apply_camera_choice(result)
            if on_chosen is not None:
                on_chosen(result)

        self._cam_dialog = CameraDialog(
            backend=_CAM_BACKEND,
            preferred_index=(self._camera_info.get("index")
                             if self._camera_info else self._load_camera_pref()),
            parent=self)
        self._cam_dialog.completed.connect(_done)
        self._cam_dialog.open()

    def _switch_camera(self):
        """Camera panel button — reopen the picker mid-session.

        Previously this cycled blindly to the next index, which gave no way to
        tell what you were switching to and could not see a webcam plugged in
        after launch.
        """
        self._open_camera_picker()

    def _on_blendshape_changed(self, name: str):
        """User picked a different blendshape in the BlendshapeWatch dropdown.
        BlendshapeWatch has already wiped its own display buffer — we just
        need to start reading a different key out of face_data each tick."""
        self._watched_blendshape = name

    def _camera_loop(self):
        while self._running:
            # The capture is None while the picker has the device, and can be
            # swapped for a different one at any moment; both are handled by
            # reading under the lock and treating "no device" as a dropped frame.
            with self._cap_lock:
                cap = self._cap
                ok, frame = cap.read() if cap is not None else (False, None)
            if ok and frame is not None and frame.mean() > 1.0:
                _cap_ns = time.time_ns()
                with self._lock:
                    self._frame_capture_ns = _cap_ns
                frame = cv2.flip(frame, 1)
                self._cam_ok = True
                with self._lock:
                    self._pending_frame = frame
                    if self._last_frame is None:
                        self._last_frame = (frame, {"detected": False})
                    else:
                        self._last_frame = (frame, self._last_frame[1])
                # Write to video file if recording is active.
                # Annotate before acquiring the writer lock (drawing is slow);
                # then check + write under the lock so release() can never
                # interleave with a write and corrupt the moov atom.
                with self._lock:
                    fd = self._last_frame[1] if self._last_frame else None
                rec_frame = self._draw_recording_overlay(
                    frame.copy(), fd, self._last_scores
                )
                with self._writer_lock:
                    if self._writer is not None:
                        try:
                            self._writer.write(rec_frame)
                            self._rec_frame_count += 1
                        except Exception:
                            pass
            else:
                # No frame this pass (camera hiccup/disconnected) — a short
                # sleep here just avoids a hot spin loop. On the success path
                # we deliberately do NOT sleep: cap.read() already blocks
                # until the camera hardware delivers its next frame, so an
                # unconditional extra sleep on top of that was silently
                # halving the real capture rate — recordings were being
                # written at roughly half the fps stamped in the file,
                # which is exactly why played-back video looked sped up.
                time.sleep(0.01)

    def _analysis_loop(self):
        # Same double-throttle bug _camera_loop had: analyze() already costs
        # real time (MediaPipe inference), so an unconditional sleep(0.033)
        # afterward was stacking on top of that and staling face_data —
        # every blendshape score, emotion, and gaze metric feeding the live
        # charts — by up to an extra frame interval on every single pass.
        # Fix: only do the (expensive) analyze() call when a genuinely new
        # frame has arrived since last time (tracked via the camera loop's
        # capture timestamp), and idle-sleep briefly otherwise. This also
        # avoids wastefully re-analyzing the same still frame back-to-back
        # now that _camera_loop can hand off frames much faster.
        last_seen_ns = 0
        while self._running:
            with self._lock:
                frame    = self._pending_frame
                frame_ns = self._frame_capture_ns
            if frame is not None and frame_ns != last_seen_ns:
                last_seen_ns = frame_ns
                small = cv2.resize(frame, (640, 360))
                face_data = self.face.analyze(small)
                with self._lock:
                    self._last_frame = (frame, face_data)
                # Raw per-frame facial logging (native rate ~30 fps)
                if self._session_start_ns and not self._session_ended:
                    fd = face_data or {}
                    self._raw_facial_rows.append({
                        "master_ts_ns": time.time_ns(),
                        "elapsed_s":    round(time.time() - self._session_start, 4),
                        "detected":     bool(fd.get("detected")),
                        "expression":   str(fd.get("dominant", "")),
                        "eye_ar":       round(float(fd.get("eye_ar") or 0), 4),
                        "blink_rate":   round(float(fd.get("blink_rate") or 0), 2),
                        "gaze_dev":     round(float(fd.get("gaze_deviation") or 0), 4),
                        "pupil_norm":   round(float(fd.get("pupil_norm") or 0), 4),
                        "duchenne":     int(fd.get("duchenne", 0)),
                    })
            else:
                # No new frame since last pass — brief idle wait rather than
                # a hot spin loop, and much shorter than the old 33ms so a
                # freshly-arrived frame gets picked up almost immediately.
                time.sleep(0.005)

    # ════════════════════════════════════════════════════════════════════════
    # Audio thread
    # ════════════════════════════════════════════════════════════════════════
    def _start_audio(self):
        # Guard against re-entry when a second session starts
        if hasattr(self, "_audio_stream") and self._audio_stream is not None:
            return
        try:
            self._sample_rate = int(sd.query_devices(kind="input")["default_samplerate"])
        except Exception:
            self._sample_rate = 44100

        def callback(indata, frames, time_info, status):
            samples = indata[:, 0].copy()
            result  = self.vocal.analyze(samples, self._sample_rate)
            n = min(len(samples), len(self._audio_buffer))
            with self._lock:
                self._last_vocal = result
                self._audio_buffer = np.roll(self._audio_buffer, -n)
                self._audio_buffer[-n:] = samples[:n]
            self._mic_ok = True
            # Raw per-chunk vocal logging (native rate ~every 0.09s at 4096/44100)
            if self._session_start_ns and not self._session_ended and result:
                self._raw_vocal_rows.append({
                    "master_ts_ns":    time.time_ns(),
                    "elapsed_s":       round(time.time() - self._session_start, 4),
                    "is_speaking":     bool(result.get("is_speaking")),
                    "pitch_stability": round(float(result.get("pitch_stability") or 0.5), 4),
                    "energy_level":    round(float(result.get("energy_level") or 0), 4),
                    "tremor_index":    round(float(result.get("tremor_index") or 0), 4),
                    "dominant_hz":     round(float(result.get("dominant_hz") or 0), 1),
                    "hnr_db":          round(float(result.get("hnr_db") or 0), 2),
                    "jitter":          round(float(result.get("jitter") or 0), 5),
                })

        try:
            self._audio_stream = sd.InputStream(channels=1, blocksize=4096, callback=callback)
            self._audio_stream.start()
        except Exception as e:
            print(f"Microphone unavailable: {e}")

    # ════════════════════════════════════════════════════════════════════════
    # Main UI tick
    # ════════════════════════════════════════════════════════════════════════
    def _update_body(self):
        # Skip when nothing dashboard-related is visible (overview / summary)
        cur = self._stack.currentWidget()
        if cur is None or cur is self._overview or cur is self._sum:
            return
        if self._session_ended:
            return

        with self._lock:
            frame_data = self._last_frame
            vocal_data = self._last_vocal
            audio_buf  = self._audio_buffer.copy()

        face_data = frame_data[1] if frame_data else None
        frame_bgr = frame_data[0] if frame_data else None

        # ── Calibration screen path ──
        if self._cal is not None and self._stack.currentWidget() is self._cal:
            # Push preview frame + indicators
            if frame_bgr is not None:
                cal_frame = frame_bgr.copy()
                self._draw_face_mesh(cal_frame, face_data)
                self._cal.update_preview(cal_frame, face_data)
            self._cal.update_indicators(
                face_detected=bool(face_data and face_data.get("detected")),
                voice_samples=len(self._calibration_vocal["pitch_stability"]),
                hrv_status=self.hrv.get_display().get("status", "disabled"),
                hrv_samples=len(self._calibration_hrv),
            )

            if self._calibrating and self._calibration_started_at is not None:
                self._cal_frames_total += 1
                if face_data and face_data.get("detected"):
                    self._cal_frames_face += 1
                self._collect_calibration_samples(face_data, vocal_data)
                elapsed = time.time() - self._calibration_started_at
                self._cal.update_progress(elapsed, self._calibration_seconds)
                if elapsed >= self._calibration_seconds:
                    self._enter_live_session()
            return

        # ── Live dashboard path ──
        # Compute scores via the trust engine
        hrv_score = self.hrv.get_score()
        scores    = self.trust.update(face_data, vocal_data, hrv_score)
        self._last_scores = scores   # Make latest scores available to the recording overlay
        pupil_now = face_data.get("pupil_norm") if face_data else None
        wl_state  = self.workload.update(pupil_now)
        self._workload_state = wl_state

        # Broadcast tick to live-stream subscribers (stub — passive only)
        _last_ev = self._event_log[-1] if self._event_log else None
        self._broadcast_tick({
            "master_ts_ns": time.time_ns(),
            "total":        scores.get("total", 50),
            "channels": {
                "facial": scores.get("facial", 50),
                "vocal":  scores.get("vocal",  50),
                "gaze":   scores.get("gaze",   50),
                "hrv":    scores.get("hrv",    65),
            },
            "dtotal":       scores.get("dscores", {}).get("total", 0.0),
            "dchannels": {
                k: scores.get("dscores", {}).get(k, 0.0)
                for k in ("facial", "vocal", "gaze", "hrv")
            },
            "active_channels": scores.get("active_channels", []),
            "last_event": _last_ev,
        })

        # Accumulate full session history (chart manages its own trailing window)
        for k, v in scores.items():
            if k not in self._history:
                continue
            self._history[k].append(v)
        self._history_t.append(time.time() - self._session_start if self._session_start else 0.0)

        # Record one row per second
        now = time.time()
        if now - self._last_record_time >= 1.0:
            self._record_row(scores, face_data, vocal_data, wl_state)
            self._last_record_time = now
            total_now = int(scores.get("total", 50))
            if self._last_flag_total is not None and (self._last_flag_total - total_now) > 10:
                self._emit_flag("Sharp trust drop", total_now)
            self._last_flag_total = total_now

        self._check_flags(scores, face_data, vocal_data)

        # ── Push to widgets ──
        if frame_bgr is not None:
            self.cam_panel.update_frame(frame_bgr, face_data)
        baseline = self._calibration_baseline if self._calibration_baseline else None
        self.cam_panel.update_metrics(face_data, baseline)

        self.score_panel.update_scores(
            scores["total"], scores["facial"], scores["vocal"],
            scores["gaze"], scores["hrv"],
        )
        self.score_panel.set_hrv_connected(self.hrv.is_connected)
        self.score_panel.update_workload(wl_state)

        # Attribution strip — 6s rolling delta (~100 ticks at 60ms)
        hist_total = self._history["total"]
        if len(hist_total) >= 6:
            delta_6s = float(hist_total[-1]) - float(hist_total[-min(100, len(hist_total))])
            self.score_panel.update_attribution(delta_6s, scores.get("contributions", {}))

        self.voice_panel.update_metrics(vocal_data, baseline)
        # Waveform downsampled, spectrum from a small FFT
        self.voice_panel.set_waveform(audio_buf[::32])
        spec = self._compute_spectrum(audio_buf)
        self.voice_panel.set_spectrum(spec)

        # Blendshape watch — raw 0-1 score for whichever blendshape is
        # currently selected in the dropdown. Not part of self._history:
        # BlendshapeWatch keeps its own short trailing window only.
        bs_value = None
        if face_data and face_data.get("detected"):
            bs_value = face_data.get("blendshapes", {}).get(self._watched_blendshape)
        self.blendshape_watch.update_value(bs_value, self._history_t[-1] if self._history_t else 0.0)

        # History chart — full session history (chart manages its own window)
        h  = {k: self._history[k] for k in ("total", "facial", "vocal", "gaze")}
        ts = self._history_t
        self.history_chart.update_traces(h, ts)
        self.history_chart.set_phases(self._phase_segments)

        # Workload glow on top strip
        if wl_state:
            self.top.setWorkloadProgress(float(wl_state.get("spike_progress", 0.0)))

        # Status dots
        self.top.set_status(
            face  = "active"  if face_data and face_data.get("detected") else "loading",
            gaze  = "active"  if face_data and face_data.get("detected") else "loading",
            voice = "active"  if vocal_data and vocal_data.get("is_speaking") else
                    "idle"    if self._mic_ok else "off",
        )

    # ════════════════════════════════════════════════════════════════════════
    # Helpers
    # ════════════════════════════════════════════════════════════════════════
    @staticmethod
    def _compute_spectrum(audio_buf: np.ndarray, n_bins: int = 48) -> list[float]:
        """Tiny FFT → log-magnitude bins, normalized 0..1."""
        if len(audio_buf) < 256:
            return [0.0] * n_bins
        # Take a power-of-two segment for FFT cleanliness
        segment = audio_buf[-1024:]
        # Windowed FFT
        window = np.hanning(len(segment))
        mag = np.abs(np.fft.rfft(segment * window))
        # Bin into log-spaced groups
        if mag.size == 0:
            return [0.0] * n_bins
        # Log-scale + normalize
        mag = np.log1p(mag)
        # Trim to lower-half spectrum (speech is sub-4k Hz typically)
        mag = mag[: max(1, len(mag) // 3)]
        # Resample down to n_bins
        idx = np.linspace(0, len(mag) - 1, n_bins).astype(int)
        bins = mag[idx]
        peak = float(bins.max()) if bins.size else 1.0
        if peak <= 0:
            return [0.0] * n_bins
        return (bins / peak).tolist()

    def _emit_flag(self, text: str, total: int):
        """Append a behavioural flag, debounced ~8s per flag type."""
        now = time.time()
        if now - self._flag_cooldowns.get(text, 0.0) < 8.0:
            return
        self._flag_cooldowns[text] = now
        color = trust_band(int(total))[1]
        ts = datetime.now().strftime("%H:%M:%S")
        self.flag_sidebar.add_flag(ts, text, color)
        self._session_flags.append((ts, text, color))

    def _check_flags(self, scores, face_data, vocal_data):
        """Real-time behavioural triggers. Reads analyzer output defensively —
        triggers whose signals aren't exposed simply never fire."""
        total = int(scores.get("total", 50))
        if face_data and face_data.get("detected"):
            if face_data.get("blink_rate", 0) > 32:
                self._emit_flag("Rapid blink rate", total)
            if face_data.get("gaze_deviation", 0) > 0.8:
                self._emit_flag("Sustained gaze aversion", total)
            aus = face_data.get("aus") or {}
            if aus.get("AU07", 0) > 0.3 and aus.get("AU04", 0) > 0.3:
                self._emit_flag("Hostile gaze detected", total)
        if vocal_data and vocal_data.get("tremor_index", 0) > 0.6:
            self._emit_flag("Voice tremor elevated", total)

    def _record_row(self, scores, face_data, vocal_data, wl_state):
        """Capture one second of data.  All %-based fields are stored
        already multiplied by 100 so the export needs no conversion."""
        now     = datetime.now()
        elapsed = round(time.time() - self._session_start, 1)

        fd = face_data or {}
        vd = vocal_data or {}
        wd = wl_state   or {}
        hd = self.hrv.get_display()

        row = {
            # ── Timestamps ─────────────────────────────────────────────────
            "timestamp":      now.strftime("%Y-%m-%d %H:%M:%S"),
            "elapsed_s":      elapsed,
            "master_ts_ns":   time.time_ns(),
            # ── Phase (researcher-marked; see TopStrip / hotkeys 1-2-3) ─────
            "phase":          (self._phase_defs[self._phase_index][1]
                                if self._phase_index >= 0 else "Unmarked"),
            # ── Composure scores ────────────────────────────────────────────
            "total":          int(scores.get("total", 50)),
            "facial":         int(scores.get("facial", 50)),
            "vocal":          int(scores.get("vocal", 50)),
            "gaze":           int(scores.get("gaze", 50)),
            "hrv":            int(scores.get("hrv", 65)),
            # ── Rate-of-change ─────────────────────────────────────────────
            "dtotal":         scores.get("dscores", {}).get("total",  0.0),
            "dfacial":        scores.get("dscores", {}).get("facial", 0.0),
            "dvocal":         scores.get("dscores", {}).get("vocal",  0.0),
            "dgaze":          scores.get("dscores", {}).get("gaze",   0.0),
            "dhrv":           scores.get("dscores", {}).get("hrv",    0.0),
            "active_channels": scores.get("active_channels", []),
            "latency_ns":     (time.time_ns() - self._frame_capture_ns) if self._frame_capture_ns else 0,
            # ── HRV (raw sensor readings from the BLE chest strap) ──────────
            # heart_rate/rmssd_ms are None until a real device is connected and streaming;
            # stored as 0 in the export rather than blank so downstream stats tools don't choke.
            "hrv_connected":  self.hrv.is_connected,
            "heart_rate_bpm": int(hd.get("heart_rate") or 0),
            "rmssd_ms":       round(float(hd.get("rmssd_ms") or 0.0), 1),
            # ── Facial ──────────────────────────────────────────────────────
            "face_det":       bool(fd.get("detected")),
            "expression":     str(fd.get("dominant", "—")),
            "eye_openness":   round(float(fd.get("eye_ar", 0)) * 100, 1),
            "blink_rate":     round(float(fd.get("blink_rate", 0)), 1),
            "gaze_dev":       round(float(fd.get("gaze_deviation", 0)) * 100, 1),
            "pupil_norm":     round(float(fd.get("pupil_norm") or 0), 4),
            "duchenne":       int(fd.get("duchenne", 0)),
            # ── Vocal ───────────────────────────────────────────────────────
            "speaking":        bool(vd.get("is_speaking")),
            "pitch_stab":      round(float(vd.get("pitch_stability", 0.5)) * 100, 1),
            "energy_level":    round(float(vd.get("energy_level",    0.0)) * 100, 1),
            "tremor":          round(float(vd.get("tremor_index",     0.0)) * 100, 1),
            "dominant_hz":     round(float(vd.get("dominant_hz",      0.0)), 1),
            # ── eGeMAPS voice-quality columns added to Excel "Vocal Analysis" sheet ──────────────
            "jitter":          round(float(vd.get("jitter",           0.0)) * 100, 4),  # stored as % (× 100) to match the panel display
            "shimmer_db":      round(float(vd.get("shimmer_db",       0.0)), 3),        # raw dB value from eGeMAPS shimmerLocaldB
            "hnr_db":          round(float(vd.get("hnr_db",           0.0)), 2),        # raw dB value from eGeMAPS HNRdBACF
            # ── eGeMAPS spectral columns added to Excel "Vocal Analysis" sheet ─────────────────
            "spectral_flux":   round(float(vd.get("spectral_flux",    0.0)), 5),        # kept at 5 dp because typical values are ~0.004–0.02
            "alpha_ratio":     round(float(vd.get("alpha_ratio",      0.0)), 3),        # negative dB value; more negative = more low-freq energy
            "hammarberg_idx":  round(float(vd.get("hammarberg_idx",   0.0)), 3),        # positive value; higher = greater vocal effort
            # ── eGeMAPS formant columns added to Excel "Vocal Analysis" sheet ──────────────────
            "f1_hz":           round(float(vd.get("f1_hz",            0.0)), 1),        # 1st formant in Hz; 0 when unvoiced
            "f2_hz":           round(float(vd.get("f2_hz",            0.0)), 1),        # 2nd formant in Hz; 0 when unvoiced
            # ── eGeMAPS MFCC columns added to Excel "Vocal Analysis" sheet ────────────────────
            "mfcc1":           round(float(vd.get("mfcc1",            0.0)), 3),
            "mfcc2":           round(float(vd.get("mfcc2",            0.0)), 3),
            "mfcc3":           round(float(vd.get("mfcc3",            0.0)), 3),
            "mfcc4":           round(float(vd.get("mfcc4",            0.0)), 3),
            # ── Cognitive load ──────────────────────────────────────────────
            "high_workload":  bool(wd.get("is_high_workload")),
            "pcps":           round(float(wd.get("pcps",           1000.0)), 2),
            "wiv":            round(float(wd.get("wiv",            1000.0)), 2),
            "spike_progress": round(float(wd.get("spike_progress",    0.0)) * 100, 1),
            # ── Action Units (approximated from MediaPipe blendshapes, 0–1) ─
            # aus dict keys: AU01 AU02 AU04 AU05 AU06 AU07 AU09 AU10
            #                AU12 AU14 AU15 AU17 AU20 AU23 AU25 AU26 AU45
            "aus":            {au: round(v, 3)
                               for au, v in fd.get("aus", {}).items()},
            # ── All 52 raw MediaPipe blendshape scores (0–1) ────────────────
            "blendshapes":    {name: round(float(v), 4)
                               for name, v in fd.get("blendshapes", {}).items()},
        }
        self._session_rows.append(row)

    def _compute_session_stats(self) -> dict:
        rows = self._session_rows
        if not rows:
            return {}
        n = len(rows)
        avg = lambda k: sum(r[k] for r in rows) / n
        pct = lambda k: 100 * sum(1 for r in rows if r[k]) / n

        durSecs = int(time.time() - self._session_start)
        durStr  = f"{durSecs // 60:02d}:{durSecs % 60:02d}"

        return {
            "duration_str":        durStr,
            "n_samples":           n,
            "trust_total":         int(round(avg("total"))),
            "trust_facial":        int(round(avg("facial"))),
            "trust_vocal":         int(round(avg("vocal"))),
            "trust_gaze":          int(round(avg("gaze"))),
            "trust_hrv":           int(round(avg("hrv"))),
            "peak_trust":          max(r["total"] for r in rows),
            "low_trust":           min(r["total"] for r in rows),
            "pct_face_detected":   pct("face_det"),
            "pct_speaking":        pct("speaking"),
            "pct_high_workload":   pct("high_workload"),
            # pitch_stab / tremor / gaze_dev already stored as %, no × 100
            "avg_pitch_stability": avg("pitch_stab"),
            "avg_tremor":          avg("tremor"),
            # eGeMAPS session averages — added alongside avg_pitch_stability and avg_tremor
            # so the session summary card can surface vocal quality trends across the whole meeting.
            "avg_hnr_db":          avg("hnr_db"),       # Mean HNR over the session; < 10 dB average suggests persistent vocal strain
            "avg_jitter":          avg("jitter"),        # Mean jitter % over the session; elevated values indicate chronic perturbation
            "avg_alpha_ratio":     avg("alpha_ratio"),   # Mean alpha ratio; trend toward 0 over a session signals increasing vocal strain
            "avg_blink_rate":      avg("blink_rate"),
            "avg_gaze_deviation":  avg("gaze_dev"),
            "trust_history":       [r["total"] for r in rows],
            "phase_segments":      list(self._phase_segments),
            "active_channels":     [ch for ch, active in self.trust._active.items() if active],
            "score_version":       SCORE_VERSION,
            "score_config":        SCORE_CONFIG,
            "flags":               list(self._session_flags),
            "participant":         dict(self._demographics),
            # Which lens this was captured through. Facial and gaze readings are
            # not comparable across a laptop camera and an external one at a
            # different height and field of view, so the device travels with the
            # session rather than being lost at the end of it.
            "camera":              dict(self._camera_info),
        }

    # ════════════════════════════════════════════════════════════════════════
    # Persistence
    # ════════════════════════════════════════════════════════════════════════
    def _load_sessions(self) -> list:
        try:
            with open(self._sessions_file, "r") as f:
                return json.load(f)
        except Exception:
            return []

    def _save_session(self, stats: dict):
        sessions = self._load_sessions()
        sessions.append({
            "date":             datetime.now().strftime("%Y-%m-%d %H:%M"),
            "session_id":       stats.get("session_id", ""),
            "duration_str":     stats.get("duration_str", "00:00"),
            "n_samples":        stats.get("n_samples", 0),
            "trust_total":      stats.get("trust_total", 50),
            "trust_facial":     stats.get("trust_facial", 50),
            "trust_vocal":      stats.get("trust_vocal", 50),
            "trust_gaze":       stats.get("trust_gaze", 50),
            "trust_hrv":        stats.get("trust_hrv", 65),
            "recording_path":   stats.get("recording_path"),   # relative or None
            "thumbnail_path":   stats.get("thumbnail_path"),   # relative or None
            "n_events":         len(self._event_log),
            "score_version":    stats.get("score_version", ""),
            "active_channels":  stats.get("active_channels", []),
            "participant":      stats.get("participant", {}),
            # Everything below is stored purely so a session card in the
            # overview list can be clicked later and show a full summary
            # (chart, phase bands, flags) without needing the original
            # in-memory session rows.
            "peak_trust":       stats.get("peak_trust", stats.get("trust_total", 50)),
            "low_trust":        stats.get("low_trust",  stats.get("trust_total", 50)),
            "trust_history":    stats.get("trust_history", []),
            "phase_segments":   stats.get("phase_segments", []),
            "flags":            stats.get("flags", []),
        })
        try:
            with open(self._sessions_file, "w") as f:
                json.dump(sessions, f, indent=2)
        except Exception as e:
            print(f"[sessions] Could not save: {e}")

    def _export_csv(self):
        if not self._session_rows:
            return
        default_name = str(
            self._session_dir / f"trust-session-{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Export session as Excel", default_name, "Excel (*.xlsx)",
        )
        if not path:
            return
        try:
            self._build_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    # ── Excel builder ────────────────────────────────────────────────────────
    def _build_excel(self, path: str):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        HDR_FILL   = PatternFill("solid", fgColor="2563EB")   # accent blue
        HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        BODY_FONT  = Font(name="Calibri", size=10)
        ALT_FILL   = PatternFill("solid", fgColor="F1F5F9")   # very light slate
        LEG_TITLE  = Font(bold=True, name="Calibri", size=10, color="1E3A5F")
        LEG_KEY    = Font(bold=True, name="Calibri", size=9)
        LEG_VAL    = Font(name="Calibri", size=9, color="475569")
        LEG_FILL   = PatternFill("solid", fgColor="EFF6FF")
        CENTER     = Alignment(horizontal="center", vertical="center")
        LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin       = Side(style="thin", color="CBD5E1")
        BORDER     = Border(bottom=thin)

        rows = self._session_rows

        def _yn(v):
            return "Yes" if v else "No"

        def _auto_width(ws, min_w=10, max_w=48):
            for col in ws.columns:
                best = min_w
                for cell in col:
                    if cell.value is not None:
                        best = max(best, min(max_w, len(str(cell.value)) + 2))
                ws.column_dimensions[get_column_letter(col[0].column)].width = best

        def _write_legend(ws, legend, n_data_rows, total_cols):
            """Shared LEGEND block writer — used by every 1/sec sheet so
            self-documentation stays consistent across the workbook."""
            leg_start = n_data_rows + 3   # blank row gap
            title_cell = ws.cell(row=leg_start, column=1, value="LEGEND")
            title_cell.font      = LEG_TITLE
            title_cell.fill      = LEG_FILL
            title_cell.alignment = LEFT
            ws.merge_cells(start_row=leg_start, start_column=1,
                           end_row=leg_start,   end_column=total_cols)

            for i, (field, desc) in enumerate(legend, leg_start + 1):
                k = ws.cell(row=i, column=1, value=field)
                k.font      = LEG_KEY
                k.alignment = LEFT
                d = ws.cell(row=i, column=2, value=desc)
                d.font      = LEG_VAL
                d.alignment = LEFT
                if total_cols > 2:
                    ws.merge_cells(start_row=i, start_column=2,
                                   end_row=i,   end_column=total_cols)

        def _write_sheet(ws, columns, legend):
            """columns: list of (header_label, row_key, formatter_fn | None)
               legend:  list of (field_name, description)"""
            # ── header ──────────────────────────────────────────────────────
            for c, (label, _, _) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=c, value=label)
                cell.font      = HDR_FONT
                cell.fill      = HDR_FILL
                cell.alignment = CENTER
                cell.border    = BORDER
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 20

            # ── data rows ───────────────────────────────────────────────────
            for r_idx, row in enumerate(rows, 2):
                fill = ALT_FILL if r_idx % 2 == 0 else None
                for c, (_, key, fmt) in enumerate(columns, 1):
                    raw = row.get(key, "")
                    val = fmt(raw) if fmt else raw
                    cell = ws.cell(row=r_idx, column=c, value=val)
                    cell.font      = BODY_FONT
                    cell.alignment = CENTER
                    if fill:
                        cell.fill = fill

            # ── legend ──────────────────────────────────────────────────────
            _write_legend(ws, legend, len(rows), len(columns))
            _auto_width(ws)

        wb = openpyxl.Workbook()

        # ════════════════════════════════════════════════════════════════════
        # Sheet 1 — Trust Session  (matches reference format exactly)
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Trust Session"
        _write_sheet(ws1, [
            ("Timestamp",        "timestamp",      None),
            ("Elapsed (s)",      "elapsed_s",      None),
            ("Phase",            "phase",          None),
            ("Trust Total",      "total",          None),
            ("Facial",           "facial",         None),
            ("Vocal",            "vocal",          None),
            ("Gaze",             "gaze",           None),
            ("HRV",              "hrv",            None),
            ("Face Detected",    "face_det",       _yn),
            ("Expression",       "expression",     None),
            ("Eye Openness %",   "eye_openness",   None),
            ("Blink Rate /min",  "blink_rate",     None),
            ("Gaze Deviation %", "gaze_dev",       None),
            ("Pupil (norm.)",    "pupil_norm",     None),
            ("Duchenne Smile",   "duchenne",       None),
            ("Speaking",         "speaking",       _yn),
            ("Pitch Stability %","pitch_stab",     None),
            ("Voice Energy %",   "energy_level",   None),
            ("Tremor Index %",   "tremor",         None),
            ("Vocal Hz",         "dominant_hz",    None),
            ("High Workload",    "high_workload",  _yn),
            ("PCPS",             "pcps",           None),
            ("WIV",              "wiv",            None),
            ("Spike Progress %", "spike_progress", None),
            ("dTotal",           "dtotal",  None),
            ("dFacial",          "dfacial", None),
            ("dVocal",           "dvocal",  None),
            ("dGaze",            "dgaze",   None),
            ("dHRV",             "dhrv",    None),
            ("Latency (ns)",     "latency_ns", None),
        ], legend=[
            ("Phase",             "Researcher-marked experiment phase in effect when this row was "
                                  "captured: Trust Establishment / Trust Violation / Trust Recovery, "
                                  "or 'Unmarked' before the researcher marks the first phase. "
                                  "Marked live via hotkeys 1/2/3 or the TopStrip button — never inferred."),
            ("Trust Total",      "Weighted composure index (0–100). "
                                 "35% Facial + 25% Vocal + 25% Gaze + 15% HRV, "
                                 "smoothed with α=0.20 exponential moving average."),
            ("Facial",           "Facial composure sub-score (0–100)."),
            ("Vocal",            "Vocal composure sub-score (0–100)."),
            ("Gaze",             "Gaze / head-pose composure sub-score (0–100)."),
            ("HRV",              "Heart-rate variability composure sub-score (0–100)."),
        ])

        # ════════════════════════════════════════════════════════════════════
        # Sheet 2 — Facial Analysis  (metrics + blendshapes)
        # ════════════════════════════════════════════════════════════════════
        # Standard columns use the shared blue header.
        # Blendshape columns are appended with a teal header so they are
        # visually distinct but still part of the same 1-fps sheet.
        HDR_FILL_BS = PatternFill("solid", fgColor="0E7490")   # teal — Blendshapes

        FACIAL_FIXED = [
            ("Timestamp",        "timestamp",    None),
            ("Elapsed (s)",      "elapsed_s",    None),
            ("Phase",            "phase",        None),
            ("Facial Score",     "facial",       None),
            ("Face Detected",    "face_det",     _yn),
            ("Expression",       "expression",   None),
            ("Eye Openness %",   "eye_openness", None),
            ("Blink Rate /min",  "blink_rate",   None),
            ("Gaze Deviation %", "gaze_dev",     None),
            ("Pupil (norm.)",    "pupil_norm",   None),
            ("Duchenne Smile",   "duchenne",     None),
        ]
        FACIAL_LEGEND = [
            ("Facial Score",     "Composure sub-score derived from eye openness stability, "
                                 "blink regularity, gaze deviation, and expression neutrality (0–100)."),
            ("Face Detected",    "Whether MediaPipe detected and tracked a face in the frame."),
            ("Expression",       "Dominant facial expression inferred from blendshape scores "
                                 "(neutral · happy · sad · angry · surprised · fearful · disgusted · contempt)."),
            ("Eye Openness %",   "Eye Aspect Ratio × 100. Typical open-eye range: 25–45 %. "
                                 "Values below ~15 % indicate a blink in progress."),
            ("Blink Rate /min",  "Rolling blinks-per-minute count. "
                                 "Normal rest: 15–20 /min. Elevated rates may indicate fatigue or stress."),
            ("Gaze Deviation %", "Head-pose deviation from camera centre (yaw + 0.5 × pitch, normalised). "
                                 "100 % = looking 40° away. Values above 20 % indicate the subject is looking away."),
            ("Pupil (norm.)",    "Iris radius normalised to inter-ocular distance via MediaPipe iris landmarks. "
                                 "Larger values indicate pupil dilation (higher cognitive arousal)."),
            ("Duchenne Smile",   "Binary flag. 1 = genuine (Duchenne) smile detected: "
                                 "AU06 (cheek raiser) active simultaneously with AU12 (lip corner puller)."),
        ]

        # Canonical MediaPipe blendshape order (ARKit topology, index 0 = _neutral)
        # — see module-level BLENDSHAPE_NAMES, shared with the live Blendshape
        # Watch panel selector.
        BS_ORDER = BLENDSHAPE_NAMES

        FULL_FACIAL_LEGEND = FACIAL_LEGEND

        ws2 = wb.create_sheet("Facial Analysis")
        n_fixed    = len(FACIAL_FIXED)
        n_bs       = len(BS_ORDER)
        total_cols = n_fixed + n_bs

        # ── header — fixed metrics (blue) ───────────────────────────────────
        for c, (hdr, _, _f) in enumerate(FACIAL_FIXED, 1):
            cell = ws2.cell(row=1, column=c, value=hdr)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER

        # ── header — blendshapes (teal) ─────────────────────────────────────
        for i, bs_name in enumerate(BS_ORDER):
            c    = n_fixed + 1 + i
            cell = ws2.cell(row=1, column=c, value=bs_name)
            cell.font = HDR_FONT; cell.fill = HDR_FILL_BS
            cell.alignment = CENTER; cell.border = BORDER

        ws2.freeze_panes = "A2"
        ws2.row_dimensions[1].height = 32

        # ── data rows ──────────────────────────────────────────────────────
        for r_idx, row in enumerate(rows, 2):
            fill = ALT_FILL if r_idx % 2 == 0 else None
            # Fixed columns
            for c, (_, key, fmt) in enumerate(FACIAL_FIXED, 1):
                raw  = row.get(key, "")
                val  = fmt(raw) if fmt else raw
                cell = ws2.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill
            # Blendshape columns
            bs_data = row.get("blendshapes", {})
            for i, bs_name in enumerate(BS_ORDER):
                c    = n_fixed + 1 + i
                val  = bs_data.get(bs_name, "")
                cell = ws2.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill

        # ── legend ─────────────────────────────────────────────────────────
        _write_legend(ws2, FULL_FACIAL_LEGEND, len(rows), total_cols)

        # ── column widths ───────────────────────────────────────────────────
        from openpyxl.utils import get_column_letter as gcl
        ws2.column_dimensions[gcl(1)].width = 22   # Timestamp
        ws2.column_dimensions[gcl(2)].width = 12   # Elapsed
        for c in range(3, n_fixed + 1):
            ws2.column_dimensions[gcl(c)].width = 16
        for c in range(n_fixed + 1, total_cols + 1):
            ws2.column_dimensions[gcl(c)].width = 18  # blendshape names are longer

        # ════════════════════════════════════════════════════════════════════
        # Sheet 3 — Vocal Analysis
        # ════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Vocal Analysis")
        _write_sheet(ws3, [
            ("Timestamp",          "timestamp",      None),
            ("Elapsed (s)",        "elapsed_s",      None),
            ("Phase",              "phase",          None),
            ("Vocal Score",        "vocal",          None),
            ("Speaking",           "speaking",       _yn),
            ("Pitch Stability %",  "pitch_stab",     None),
            ("Voice Energy %",     "energy_level",   None),
            ("Tremor Index %",     "tremor",         None),
            ("Vocal Hz",           "dominant_hz",    None),
            ("Jitter %",           "jitter",         None),
            ("Shimmer (dB)",       "shimmer_db",     None),
            ("HNR (dB)",           "hnr_db",         None),
            ("Spectral Flux",      "spectral_flux",  None),
            ("Alpha Ratio",        "alpha_ratio",    None),
            ("Hammarberg Index",   "hammarberg_idx", None),
            ("F1 (Hz)",            "f1_hz",          None),
            ("F2 (Hz)",            "f2_hz",          None),
            ("MFCC 1",             "mfcc1",          None),
            ("MFCC 2",             "mfcc2",          None),
            ("MFCC 3",             "mfcc3",          None),
            ("MFCC 4",             "mfcc4",          None),
        ], legend=[
            ("Vocal Score",        "Composure sub-score derived from pitch stability, energy, "
                                   "tremor, alpha ratio, and spectral flux (0–100). "
                                   "Features extracted via eGeMAPSv02 (OpenSMILE) when available."),
            ("Speaking",           "Active speech detected: perceptual loudness above the silence threshold."),
            ("Pitch Stability %",  "Inverse coefficient of variation of eGeMAPS F0 over a 60-frame history. "
                                   "100 % = perfectly stable pitch. Low values suggest vocal stress or emotion."),
            ("Voice Energy %",     "Perceptual loudness (eGeMAPS Loudness_sma3) normalised to 0–100 %."),
            ("Tremor Index %",     "Composite vocal instability: 40 % jitter + 40 % shimmer + 20 % inverted HNR. "
                                   "Values above 30 % suggest significant vocal tremor or strain."),
            ("Vocal Hz",           "Fundamental frequency (F0) in Hz converted from eGeMAPS semitones. "
                                   "Typical speech: 80–450 Hz. 0 = not speaking or unvoiced frame."),
            ("Jitter %",           "Local jitter × 100: cycle-to-cycle F0 perturbation in voiced frames. "
                                   "Normal speech < 1 %. Values above 2 % indicate vocal instability."),
            ("Shimmer (dB)",       "Local shimmer in dB: cycle-to-cycle amplitude perturbation. "
                                   "Normal speech < 1 dB. Values above 2 dB indicate vocal strain."),
            ("HNR (dB)",           "Harmonics-to-Noise Ratio. Normal speech > 20 dB. "
                                   "< 10 dB indicates a noisy, tense, or fatigued voice."),
            ("Spectral Flux",      "Mean frame-to-frame spectral change. Higher values indicate "
                                   "rapid vocal instability or agitation."),
            ("Alpha Ratio",        "Log ratio of energy in 1–5 kHz vs 50 Hz–1 kHz bands. "
                                   "More negative = energy in low frequencies (normal). "
                                   "Less negative = high-frequency dominant (strained or breathy voice)."),
            ("Hammarberg Index",   "Strongest energy peak in 2–5 kHz relative to energy below 2 kHz. "
                                   "Higher values indicate greater vocal effort and brightness."),
            ("F1 (Hz)",            "First formant frequency. Reflects vowel openness and jaw position. "
                                   "Typical range: 300–900 Hz in conversational speech."),
            ("F2 (Hz)",            "Second formant frequency. Reflects front/back vowel articulation. "
                                   "Typical range: 800–2500 Hz."),
            ("MFCC 1–4",           "Mel-Frequency Cepstral Coefficients 1–4. Encode vocal tract shape "
                                   "and timbre. Used in ML-based emotion and stress classification."),
        ])

        # ════════════════════════════════════════════════════════════════════
        # Sheet 4 — Gaze Analysis
        # ════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Gaze Analysis")
        _write_sheet(ws4, [
            ("Timestamp",          "timestamp", None),
            ("Elapsed (s)",        "elapsed_s", None),
            ("Phase",              "phase",     None),
            ("Gaze Score",         "gaze",      None),
            ("Gaze Deviation %",   "gaze_dev",  None),
            ("Pupil (norm.)",      "pupil_norm",None),
        ], legend=[
            ("Gaze Score",         "Composure sub-score based on sustained head-pose stability (0–100). "
                                   "Frequent or large deviations lower the score."),
            ("Gaze Deviation %",   "Angular deviation computed from MediaPipe 3-D head-pose landmarks: "
                                   "(|yaw| + 0.5 × |pitch|) ÷ 40°, clamped to 100 %. "
                                   "Values above 20–25 % typically indicate deliberate look-away."),
            ("Pupil (norm.)",      "Included here because iris size encodes arousal "
                                   "and correlates with sustained attention."),
        ])

        # ════════════════════════════════════════════════════════════════════
        # Sheet 5 — Cognitive Load
        # ════════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet("Cognitive Load")
        _write_sheet(ws5, [
            ("Timestamp",          "timestamp",      None),
            ("Elapsed (s)",        "elapsed_s",      None),
            ("Phase",              "phase",          None),
            ("High Workload",      "high_workload",  _yn),
            ("PCPS",               "pcps",           None),
            ("WIV",                "wiv",            None),
            ("Spike Progress %",   "spike_progress", None),
        ], legend=[
            ("PCPS",               "Pupil Change Per Second — baseline-corrected real-time pupil dilation. "
                                   "Baseline pupil = 1000. Values > 1000 indicate dilation above baseline."),
            ("WIV",                "Within-session Inertia Value — 60-second rolling mean of PCPS. "
                                   "Adapts to the subject's typical pupil level throughout the session."),
            ("High Workload",      "'Yes' when PCPS > WIV × 1.015 "
                                   "(pupil at least 1.5 % above the rolling average)."),
            ("Spike Progress %",   "Progress (0–100 %) toward a confirmed high-workload spike. "
                                   "A spike is declared when High Workload persists for 60 continuous seconds. "
                                   "Resets to 0 % on any low-workload moment."),
        ])

        # ════════════════════════════════════════════════════════════════════
        # Sheet 6 — HRV
        # ════════════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet("HRV")
        _write_sheet(ws6, [
            ("Timestamp",          "timestamp",      None),
            ("Elapsed (s)",        "elapsed_s",      None),
            ("Phase",              "phase",          None),
            ("HRV Score",          "hrv",            None),
            ("Sensor Connected",   "hrv_connected",  _yn),
            ("Heart Rate (bpm)",   "heart_rate_bpm", None),
            ("RMSSD (ms)",         "rmssd_ms",       None),
        ], legend=[
            ("HRV Score",          "Heart-rate variability composure sub-score (0–100), derived from RMSSD "
                                   "via a literature-based mapping (20-90 range). Falls back to a fixed "
                                   "stub value (65) whenever the sensor is disconnected or hasn't yet "
                                   "collected enough beats for a reading — check 'Sensor Connected' and "
                                   "'RMSSD (ms)' to tell real readings from the fallback."),
            ("Sensor Connected",   "Whether the BLE heart-rate strap (e.g. Polar H10) was connected and "
                                   "streaming at this sample."),
            ("Heart Rate (bpm)",   "Instantaneous heart rate from the most recent BLE notification. "
                                   "0 when no sensor is connected."),
            ("RMSSD (ms)",         "Root mean square of successive R-R interval differences, over a "
                                   "rolling 60-second window — the standard time-domain HRV metric that "
                                   "the HRV Score is derived from. 0 until at least 4 R-R intervals have "
                                   "been collected in the window."),
        ])


        # ═══════════════════════════════════════════════════════════════════════════
        # Sheet — Raw Facial (~30 fps)
        # ═══════════════════════════════════════════════════════════════════════════
        if self._raw_facial_rows:
            ws_rf = wb.create_sheet("Raw Facial")
            rf_cols = [
                ("Master TS (ns)", "master_ts_ns", None),
                ("Elapsed (s)",    "elapsed_s",   None),
                ("Detected",       "detected",    _yn),
                ("Expression",     "expression",  None),
                ("Eye AR",         "eye_ar",      None),
                ("Blink /min",     "blink_rate",  None),
                ("Gaze Dev",       "gaze_dev",    None),
                ("Pupil (norm)",   "pupil_norm",  None),
                ("Duchenne",       "duchenne",    None),
            ]
            for c, (hdr, _, _f) in enumerate(rf_cols, 1):
                cell = ws_rf.cell(row=1, column=c, value=hdr)
                cell.font = HDR_FONT; cell.fill = HDR_FILL
                cell.alignment = CENTER; cell.border = BORDER
            ws_rf.freeze_panes = "A2"
            for r_idx, rrow in enumerate(self._raw_facial_rows, 2):
                fill = ALT_FILL if r_idx % 2 == 0 else None
                for c, (_, key, fmt) in enumerate(rf_cols, 1):
                    raw = rrow.get(key, "")
                    val = fmt(raw) if fmt else raw
                    cell = ws_rf.cell(row=r_idx, column=c, value=val)
                    cell.font = BODY_FONT; cell.alignment = CENTER
                    if fill: cell.fill = fill
            _write_legend(ws_rf, [
                ("Master TS (ns)", "Shared nanosecond clock used to align this sheet with "
                                    "Raw Vocal and the 1/sec sheets despite differing sample rates."),
                ("Eye AR",          "Raw Eye Aspect Ratio (0–1) before the ×100 scaling used on the 1/sec sheets."),
                ("Blink /min",      "Rolling blinks-per-minute count, sampled at camera frame rate (~30 fps) "
                                    "rather than the 1/sec throttle used elsewhere."),
                ("Gaze Dev",        "Raw head-pose deviation (0–1) before the ×100 scaling used on the 1/sec sheets."),
                ("Pupil (norm)",    "Iris radius normalised to inter-ocular distance, sampled every frame."),
                ("Duchenne",        "Binary flag: AU06 + AU12 both active in this frame."),
            ], len(self._raw_facial_rows), len(rf_cols))
            _auto_width(ws_rf)

        # ═══════════════════════════════════════════════════════════════════════════
        # Sheet — Raw Vocal (per audio chunk)
        # ═══════════════════════════════════════════════════════════════════════════
        if self._raw_vocal_rows:
            ws_rv = wb.create_sheet("Raw Vocal")
            rv_cols = [
                ("Master TS (ns)",   "master_ts_ns",    None),
                ("Elapsed (s)",      "elapsed_s",       None),
                ("Speaking",         "is_speaking",     _yn),
                ("Pitch Stability",  "pitch_stability", None),
                ("Energy Level",     "energy_level",    None),
                ("Tremor Index",     "tremor_index",    None),
                ("Vocal Hz",         "dominant_hz",     None),
                ("HNR (dB)",         "hnr_db",          None),
                ("Jitter",           "jitter",          None),
            ]
            for c, (hdr, _, _f) in enumerate(rv_cols, 1):
                cell = ws_rv.cell(row=1, column=c, value=hdr)
                cell.font = HDR_FONT; cell.fill = HDR_FILL
                cell.alignment = CENTER; cell.border = BORDER
            ws_rv.freeze_panes = "A2"
            for r_idx, rrow in enumerate(self._raw_vocal_rows, 2):
                fill = ALT_FILL if r_idx % 2 == 0 else None
                for c, (_, key, fmt) in enumerate(rv_cols, 1):
                    raw = rrow.get(key, "")
                    val = fmt(raw) if fmt else raw
                    cell = ws_rv.cell(row=r_idx, column=c, value=val)
                    cell.font = BODY_FONT; cell.alignment = CENTER
                    if fill: cell.fill = fill
            _write_legend(ws_rv, [
                ("Master TS (ns)",   "Shared nanosecond clock used to align this sheet with "
                                      "Raw Facial and the 1/sec sheets despite differing sample rates."),
                ("Pitch Stability",  "Raw 0–1 value before the ×100 scaling used on the Vocal Analysis sheet."),
                ("Energy Level",     "Raw 0–1 perceptual loudness before the ×100 scaling used elsewhere."),
                ("Tremor Index",     "Raw 0–1 composite vocal instability, sampled once per audio chunk "
                                      "rather than the 1/sec throttle used on the Vocal Analysis sheet."),
                ("HNR (dB)",         "Harmonics-to-Noise Ratio for this audio chunk. Normal speech > 20 dB."),
                ("Jitter",           "Local jitter (fraction, not ×100) for this audio chunk. Normal speech < 0.01."),
            ], len(self._raw_vocal_rows), len(rv_cols))
            _auto_width(ws_rv)

        # ═══════════════════════════════════════════════════════════════════════════
        # Sheet — Score Config
        # ═══════════════════════════════════════════════════════════════════════════
        ws_cfg = wb.create_sheet("Score Config")
        cfg_rows = []
        def _flatten_cfg(d, prefix=""):
            for k, v in d.items():
                full_key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
                if isinstance(v, dict):
                    _flatten_cfg(v, full_key)
                else:
                    cfg_rows.append((full_key, str(v)))
        _flatten_cfg(SCORE_CONFIG)
        # The config above is the same for every session; what actually varies
        # per participant is the baseline measured during their calibration
        # window, so record it here — without it a session's numbers cannot be
        # reproduced or compared against another participant's.
        _flatten_cfg(self.trust.input_baseline, "measured_baseline")
        cfg_rows.append(("measured_baseline.rmssd_ms",
                         str(self.hrv.get_display().get("baseline_rmssd"))))
        for ch, val in self.trust.baseline.items():
            cfg_rows.append((f"measured_baseline.resting_score.{ch}", str(round(val, 2))))
        for c, hdr in enumerate(["Parameter", "Value"], 1):
            cell = ws_cfg.cell(row=1, column=c, value=hdr)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER
        ws_cfg.freeze_panes = "A2"
        for r_idx, (k, v) in enumerate(cfg_rows, 2):
            fill = ALT_FILL if r_idx % 2 == 0 else None
            for c, val in enumerate([k, v], 1):
                cell = ws_cfg.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill
        _auto_width(ws_cfg)

        # ═══════════════════════════════════════════════════════════════════════════
        # Sheet — Events
        # ═══════════════════════════════════════════════════════════════════════════
        ws_ev = wb.create_sheet("Events")
        ev_cols = [
            ("Master TS (ns)",  "master_ts_ns",  None),
            ("Elapsed (s)",     "elapsed_s",     None),
            ("Kind",            "kind",          None),
            ("Label",           "label",         None),
            ("Wall Clock",      "wall_clock",    None),
        ]
        for c, (hdr, _, _f) in enumerate(ev_cols, 1):
            cell = ws_ev.cell(row=1, column=c, value=hdr)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER
        ws_ev.freeze_panes = "A2"
        for r_idx, ev in enumerate(self._event_log, 2):
            fill = ALT_FILL if r_idx % 2 == 0 else None
            for c, (_, key, fmt) in enumerate(ev_cols, 1):
                raw = ev.get(key, "")
                val = fmt(raw) if fmt else raw
                cell = ws_ev.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill
        _auto_width(ws_ev)

        # ═══════════════════════════════════════════════════════════════════════════
        # Sheet — Flags (behavioural triggers, same log-sheet pattern as Events)
        # ═══════════════════════════════════════════════════════════════════════════
        ws_fl = wb.create_sheet("Flags")
        fl_cols = ["Time", "Flag", "Trust Band"]
        band_label_by_color = {color: label for _, label, color in TRUST_BANDS}
        for c, hdr in enumerate(fl_cols, 1):
            cell = ws_fl.cell(row=1, column=c, value=hdr)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER
        ws_fl.freeze_panes = "A2"
        for r_idx, (ts, text, color) in enumerate(self._session_flags, 2):
            fill = ALT_FILL if r_idx % 2 == 0 else None
            for c, val in enumerate([ts, text, band_label_by_color.get(color, color)], 1):
                cell = ws_fl.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill
            band_cell = ws_fl.cell(row=r_idx, column=3)
            band_cell.fill = PatternFill("solid", fgColor=color.lstrip("#"))
        _write_legend(ws_fl, [
            ("Flag",        "Behavioural trigger detected live during the session "
                             "(rapid blink rate, sustained gaze aversion, hostile-gaze AU combination, "
                             "voice tremor, sharp trust drop). Each flag type is debounced ~8s."),
            ("Trust Band",  "Trust-total band in effect at the moment the flag fired, "
                             "matching the colour bands used in the live sidebar."),
        ], len(self._session_flags), len(fl_cols))
        _auto_width(ws_fl)

        # ═══════════════════════════════════════════════════════════════════════════
        # Sheet — Summary (session stats + phase breakdown, moved to front of workbook)
        # ═══════════════════════════════════════════════════════════════════════════
        stats = self._compute_session_stats()
        ws_sum = wb.create_sheet("Summary")
        for c, hdr in enumerate(["Metric", "Value"], 1):
            cell = ws_sum.cell(row=1, column=c, value=hdr)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER
        ws_sum.freeze_panes = "A2"

        def _r1(v):
            return round(v, 1) if isinstance(v, float) else v

        participant = stats.get("participant", {}) or {}
        camera = stats.get("camera", {}) or {}
        summary_rows = [
            ("Session ID",              getattr(self, "_session_id", "")),
            ("Participant Sex",         participant.get("sex", "—")),
            ("Participant Age",         participant.get("age", "—")),
            ("Participant Culture",     participant.get("culture") or "—"),
            ("Score Engine Version",    stats.get("score_version", "")),
            ("Camera",                  (camera.get("name") or "—")),
            ("Camera Connection",       (camera.get("label")
                                         or camera.get("transport") or "—")),
            ("Active Channels",         ", ".join(stats.get("active_channels", [])) or "—"),
            ("Duration",                stats.get("duration_str", "")),
            ("Samples Recorded",        stats.get("n_samples", 0)),
            ("Behavioural Flags Triggered", len(self._session_flags)),
            ("Avg Trust Total",         _r1(stats.get("trust_total"))),
            ("Peak Trust",              stats.get("peak_trust")),
            ("Low Trust",               stats.get("low_trust")),
            ("Avg Facial",              _r1(stats.get("trust_facial"))),
            ("Avg Vocal",               _r1(stats.get("trust_vocal"))),
            ("Avg Gaze",                _r1(stats.get("trust_gaze"))),
            ("Avg HRV",                 _r1(stats.get("trust_hrv"))),
            ("Face Detected %",         _r1(stats.get("pct_face_detected"))),
            ("Speaking %",              _r1(stats.get("pct_speaking"))),
            ("High Workload %",         _r1(stats.get("pct_high_workload"))),
            ("Avg Pitch Stability %",   _r1(stats.get("avg_pitch_stability"))),
            ("Avg Tremor %",            _r1(stats.get("avg_tremor"))),
            ("Avg HNR (dB)",            _r1(stats.get("avg_hnr_db"))),
            ("Avg Jitter %",            _r1(stats.get("avg_jitter"))),
            ("Avg Alpha Ratio",         _r1(stats.get("avg_alpha_ratio"))),
            ("Avg Blink Rate /min",     _r1(stats.get("avg_blink_rate"))),
            ("Avg Gaze Deviation %",    _r1(stats.get("avg_gaze_deviation"))),
        ]
        for r_idx, (k, v) in enumerate(summary_rows, 2):
            fill = ALT_FILL if r_idx % 2 == 0 else None
            for c, val in enumerate([k, v], 1):
                cell = ws_sum.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill

        # ── phase breakdown table ────────────────────────────────────────────
        ph_title_row = len(summary_rows) + 4
        ph_title = ws_sum.cell(row=ph_title_row, column=1, value="PHASE BREAKDOWN")
        ph_title.font = LEG_TITLE; ph_title.fill = LEG_FILL; ph_title.alignment = LEFT
        ws_sum.merge_cells(start_row=ph_title_row, start_column=1,
                            end_row=ph_title_row,   end_column=2)

        ph_hdr_row = ph_title_row + 1
        ph_cols = ["Phase", "Start (s)", "End (s)", "Duration (s)"]
        for c, hdr in enumerate(ph_cols, 1):
            cell = ws_sum.cell(row=ph_hdr_row, column=c, value=hdr)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CENTER; cell.border = BORDER

        last_elapsed = rows[-1]["elapsed_s"] if rows else 0
        for i, seg in enumerate(self._phase_segments):
            r_idx = ph_hdr_row + 1 + i
            fill = ALT_FILL if r_idx % 2 == 0 else None
            start_s = seg.get("start_s", 0)
            end_s   = seg.get("end_s")
            if end_s is None:
                end_s = last_elapsed   # phase still open when the session ended
            duration = round(end_s - start_s, 1)
            for c, val in enumerate([seg.get("label", seg.get("key", "")),
                                      round(start_s, 1), round(end_s, 1), duration], 1):
                cell = ws_sum.cell(row=r_idx, column=c, value=val)
                cell.font = BODY_FONT; cell.alignment = CENTER
                if fill: cell.fill = fill

        _auto_width(ws_sum)

        # Summary is the most useful landing page for someone opening the file
        # cold — move it to the front without disturbing the build order above.
        wb.move_sheet("Summary", offset=-wb.sheetnames.index("Summary"))

        wb.save(path)

    # ════════════════════════════════════════════════════════════════════════
    # Event logging + sync flash
    # ════════════════════════════════════════════════════════════════════════
    def log_event(self, kind: str, label: str):
        """Log a timestamped event. kind ∈ {'sync', 'breach', 'control', 'manual'}."""
        if not self._session_start_ns:
            return
        entry = {
            "master_ts_ns": time.time_ns(),
            "elapsed_s":    round(time.time() - self._session_start, 3),
            "kind":         kind,
            "label":        label,
            "wall_clock":   datetime.now().isoformat(),
        }
        self._event_log.append(entry)
        print(f"[event] {kind}/{label}  elapsed={entry['elapsed_s']}s", flush=True)

    # ════════════════════════════════════════════════════════════════════════
    # Phase tracking (researcher-marked, not inferred from the score)
    # ════════════════════════════════════════════════════════════════════════
    def _on_phase_button_clicked(self):
        """TopStrip 'Mark: <next phase>' button — advance to the next phase."""
        if self._phase_index + 1 < len(self._phase_defs):
            self._advance_phase(self._phase_index + 1)

    def _advance_phase(self, target_index: int):
        """Mark the start of a phase. Enforced sequential order: target_index
        must be exactly one past the current phase (so the researcher can't
        skip Violation, or re-mark Establishment after Recovery has begun)."""
        if not self._session_start_ns or self._session_ended:
            return
        if target_index != self._phase_index + 1:
            if target_index <= self._phase_index:
                label = self._phase_defs[target_index][1]
                self._show_sync_flash(f"{label.upper()} ALREADY MARKED", color="#8a91a1", prefix="◆ ")
            else:
                next_label = self._phase_defs[self._phase_index + 1][1]
                self._show_sync_flash(f"MARK {next_label.upper()} FIRST", color=DANGER, prefix="◆ ")
            return

        key, label, color = self._phase_defs[target_index]
        now_elapsed = round(time.time() - self._session_start, 3)
        if self._phase_segments:
            self._phase_segments[-1]["end_s"] = now_elapsed
        self._phase_segments.append({
            "key": key, "label": label, "color": color,
            "start_s": now_elapsed, "end_s": None,
        })
        self._phase_index = target_index
        self.log_event("phase", f"{key}_start")
        self._show_sync_flash(label.upper(), color=color, prefix="◆ PHASE: ")
        self.top.set_phase(label, color, target_index, len(self._phase_defs))

    def _show_sync_flash(self, text: str, color: str = "#facc15", prefix: str = "⬤ SYNC: "):
        """Show a brief on-screen marker flash. Used for WorldCam sync markers
        (default yellow) and, with a custom color/prefix, for phase-transition
        feedback."""
        from PyQt6.QtWidgets import QLabel
        from PyQt6.QtCore import Qt
        text_color = "#000" if color == "#facc15" else "#fff"
        flash = QLabel(f"{prefix}{text}", self)
        flash.setStyleSheet(
            f"background: {color}; color: {text_color}; font-size: 18px; font-weight: bold;"
            "padding: 8px 18px; border-radius: 6px;"
        )
        flash.setAlignment(Qt.AlignmentFlag.AlignCenter)
        flash.adjustSize()
        # Position top-centre
        flash.move((self.width() - flash.width()) // 2, 18)
        flash.raise_()
        flash.show()
        QTimer.singleShot(1500, flash.deleteLater)

    def keyPressEvent(self, event):
        key = event.key()
        # Only fire hotkeys during a live session
        if self._session_start_ns and not self._session_ended:
            if key == Qt.Key.Key_B:
                self.log_event("breach", "manual_breach")
                self._show_sync_flash("BREACH")
            elif key == Qt.Key.Key_C:
                self.log_event("control", "manual_control")
                self._show_sync_flash("CONTROL")
            elif key == Qt.Key.Key_M:
                self.log_event("manual", "manual_marker")
                self._show_sync_flash("MARKER")
            elif key == Qt.Key.Key_1:
                self._advance_phase(0)   # Mark: Trust Establishment
            elif key == Qt.Key.Key_2:
                self._advance_phase(1)   # Mark: Trust Violation
            elif key == Qt.Key.Key_3:
                self._advance_phase(2)   # Mark: Trust Recovery
        super().keyPressEvent(event)

    # ════════════════════════════════════════════════════════════════════════
    # WebSocket live stream (stub)
    # ════════════════════════════════════════════════════════════════════════
    def _start_ws_server(self):
        if not _HAS_WS or not self._ws_enabled:
            return
        import asyncio

        async def _handler(ws):
            self._ws_clients.add(ws)
            try:
                await ws.wait_closed()
            finally:
                self._ws_clients.discard(ws)

        async def _broadcaster():
            while True:
                payload = await self._ws_queue.get()
                dead = set()
                for ws in list(self._ws_clients):
                    try:
                        await ws.send(payload)
                    except Exception:
                        dead.add(ws)
                self._ws_clients -= dead

        async def _run():
            self._ws_queue = asyncio.Queue()
            server = await websockets.serve(_handler, "127.0.0.1", 8765)
            asyncio.create_task(_broadcaster())
            print("[ws] Live stream on ws://127.0.0.1:8765", flush=True)
            await server.wait_closed()

        def _thread():
            loop = asyncio.new_event_loop()
            self._ws_loop = loop
            loop.run_until_complete(_run())

        threading.Thread(target=_thread, daemon=True).start()

    def _broadcast_tick(self, payload_dict: dict):
        if not _HAS_WS or self._ws_loop is None or self._ws_queue is None:
            return
        if not self._ws_clients:
            return
        try:
            msg = json.dumps(payload_dict)
            self._ws_loop.call_soon_threadsafe(self._ws_queue.put_nowait, msg)
        except Exception:
            pass

    # ════════════════════════════════════════════════════════════════════════
    # Window lifecycle
    # ════════════════════════════════════════════════════════════════════════
    def closeEvent(self, event):
        self._running = False
        self.hrv.stop()
        # Flush and release writer before anything else
        with self._writer_lock:
            writer = self._writer
            self._writer = None
        if writer is not None:
            try:
                writer.release()
            except Exception:
                pass
        try:
            self._release_capture()
        except Exception:
            pass
        try:
            if hasattr(self, "_audio_stream"):
                self._audio_stream.stop()
                self._audio_stream.close()
        except Exception:
            pass
        super().closeEvent(event)


# ═══════════════════════════════════════════════════════════════════════════
def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Trust")
    app.setOrganizationName("Trust")

    # Measure the screen and scale the design to fit it. Must run before any
    # widget is constructed — sizes are read once, at construction.
    scale = init_ui_scale(app)
    geo = app.primaryScreen().availableGeometry()
    print(f"[ui] screen {geo.width()}x{geo.height()} logical px, "
          f"devicePixelRatio {app.primaryScreen().devicePixelRatio():.2f}, "
          f"UI scale {scale:.2f}", flush=True)

    # Ship Inter / JetBrainsMono .ttf files in a fonts/ folder for pixel parity
    # with the design preview. Falls back to system fonts if not present.
    load_packaged_fonts()
    app.setFont(ui_font(10))

    # ── Global QSS: neutralise macOS native chrome that bleeds through ────────
    app.setStyleSheet(f"""
        /* Slim, on-brand scrollbars */
        QScrollBar:vertical {{
            background: {BG_DEEP}; width: 8px; margin: 0; border: none;
        }}
        QScrollBar::handle:vertical {{
            background: {LINE}; border-radius: 4px; min-height: 28px;
        }}
        QScrollBar::handle:vertical:hover  {{ background: {TEXT_FAINT}; }}
        QScrollBar::add-line:vertical,
        QScrollBar::sub-line:vertical      {{ height: 0; border: none; }}
        QScrollBar::add-page:vertical,
        QScrollBar::sub-page:vertical      {{ background: none; }}

        QScrollBar:horizontal {{
            background: {BG_DEEP}; height: 8px; margin: 0; border: none;
        }}
        QScrollBar::handle:horizontal {{
            background: {LINE}; border-radius: 4px; min-width: 28px;
        }}
        QScrollBar::handle:horizontal:hover {{ background: {TEXT_FAINT}; }}
        QScrollBar::add-line:horizontal,
        QScrollBar::sub-line:horizontal    {{ width: 0; border: none; }}
        QScrollBar::add-page:horizontal,
        QScrollBar::sub-page:horizontal    {{ background: none; }}

        /* Tooltip */
        QToolTip {{
            background: {PANEL}; color: {TEXT};
            border: 1px solid {LINE}; border-radius: 4px;
            padding: 4px 8px;
        }}

        /* Suppress the native macOS focus rectangle on buttons */
        QPushButton:focus {{ outline: none; }}
        QPushButton {{ outline: none; }}
    """)

    # Allow Ctrl+C to quit: Qt's event loop blocks Python signal delivery,
    # so a no-op timer forces the interpreter to wake up periodically and
    # check for pending signals.
    signal.signal(signal.SIGINT, lambda *_: app.quit())
    _sigint_timer = QTimer()
    _sigint_timer.start(200)
    _sigint_timer.timeout.connect(lambda: None)

    w = TrustDashboard()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
